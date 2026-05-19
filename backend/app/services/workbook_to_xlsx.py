"""
Workbook JSON → xlsx renderer.

Copies the longevity template xlsx to backend/data/ on first call, then rebuilds
the editable sheets from the canonical ``plan_json`` so a downloaded plan always
reflects the latest edited tasks / rotation / screenings — not the static template.

``Plan.plan_json`` is the single source of truth (MODIFY_WORKSHEET_PLAN_FINAL
§1, §4.2). ``generate_xlsx`` stamps ``01_Personal_Settings`` and then, for each
of the brief-today, supplements, two screening, and 30-day rotation sheets,
clears the existing data rows and repopulates them from the JSON arrays, writing
into the SAME columns the deterministic ingest readers in ``plan_ingest.py``
read from (round-trip consistency).

Resilience (§11 "xlsx template sheet/column shape differs"): each sheet's
rewrite is wrapped in its own try/except. A missing sheet or a header/column
shape that does not match the expected layout skips THAT sheet silently (logged
via ``print`` like the rest of the module) and leaves it as the template — one
bad sheet never aborts the others or the download.
"""
import io
import re
import shutil
from pathlib import Path
from typing import Any

import openpyxl

# backend/data/template.xlsx — local copy used at runtime
TEMPLATE_PATH = Path(__file__).parent.parent.parent / "data" / "template.xlsx"

# Source: the v5 workbook at the project root
SOURCE_TEMPLATE = (
    Path(__file__).parent.parent.parent.parent / "Longevity_Master_together_OS_v5.xlsx"
)

# Supplement status values that mark a real task row (mirrors the set used by
# plan_ingest._read_supplements_from_excel). Anything else below the header is a
# different section (stacking rules / algorithm) and bounds the data block.
_SUPPLEMENT_STATUSES = {
    "active",
    "discuss",
    "caution",
    "avoid",
    "deprioritize",
    "optional",
    "specialist only",
}


def get_or_copy_template() -> Path:
    """Ensure backend/data/template.xlsx exists; copy from source if needed."""
    TEMPLATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    if not TEMPLATE_PATH.exists() and SOURCE_TEMPLATE.exists():
        shutil.copy2(SOURCE_TEMPLATE, TEMPLATE_PATH)
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            f"Template xlsx not found. Expected at {TEMPLATE_PATH} "
            f"(source: {SOURCE_TEMPLATE})."
        )
    return TEMPLATE_PATH


# ── normalization (mirrors plan_ingest._normalize_pillar) ──────────────────


def _normalize_pillar(sheet_name: str) -> str:
    name = re.sub(r"^\d+_?", "", sheet_name.strip())
    return name.lower().replace(" ", "_")


def _find_sheet(wb: openpyxl.Workbook, pillar: str) -> Any:
    """Return the worksheet whose normalized name matches *pillar*, or None."""
    for sheet_name in wb.sheetnames:
        if _normalize_pillar(sheet_name) == pillar:
            return wb[sheet_name]
    return None


def _row_text(ws: Any, row_idx: int) -> list[str]:
    """1-indexed row → list of trimmed string cell values."""
    out: list[str] = []
    for cell in ws[row_idx]:
        out.append("" if cell.value is None else str(cell.value).strip())
    return out


def _find_header_row(ws: Any, required: list[str], limit: int = 25) -> int | None:
    """
    Return the 1-indexed row whose cells (lowercased) contain every keyword in
    *required* as a substring. Mirrors the readers' header detection.
    """
    for r in range(1, min(ws.max_row, limit) + 1):
        lowered = [c.lower() for c in _row_text(ws, r)]
        if all(any(req in cell for cell in lowered) for req in required):
            return r
    return None


def _col_map(ws: Any, header_row: int) -> dict[str, int]:
    """Header row → {lowercased header text: 1-indexed column}."""
    cols: dict[str, int] = {}
    for cell in ws[header_row]:
        if cell.value is not None:
            text = str(cell.value).strip().lower()
            if text and text not in cols:
                cols[text] = cell.column
    return cols


def _find_col(cols: dict[str, int], *keywords: str) -> int | None:
    for kw in keywords:
        for header, idx in cols.items():
            if kw in header:
                return idx
    return None


def _clear_block(ws: Any, first_row: int, last_row: int, columns: list[int]) -> None:
    """Blank the given columns for rows [first_row, last_row] inclusive."""
    for r in range(first_row, last_row + 1):
        for c in columns:
            ws.cell(row=r, column=c).value = None


def _meta(task: dict, key: str) -> Any:
    meta = task.get("extra_metadata")
    if isinstance(meta, dict):
        return meta.get(key)
    return None


# ── per-sheet writers (each raises on shape mismatch; caller skips) ─────────


def _write_brief_today(ws: Any, tasks: list[dict]) -> None:
    """
    Rebuild the 02_Brief_Today main exercise table AND the weekly-minimums
    table from brief_today tasks, inverting _read_brief_today_from_excel so the
    round-trip is lossless.
    """
    actionable = [
        t
        for t in tasks
        if t.get("pillar") == "brief_today" and not t.get("is_reference")
    ]
    main_tasks = [t for t in actionable if not str(t.get("name", "")).startswith("(must)")]
    must_tasks = [t for t in actionable if str(t.get("name", "")).startswith("(must)")]

    # --- main exercise table ---
    main_hdr = _find_header_row(ws, ["start time", "exercise 1", "dose"])
    if main_hdr is None:
        raise ValueError("02_Brief_Today: main exercise header not found")
    cols = _col_map(ws, main_hdr)

    c_start = _find_col(cols, "start time")
    c_end = _find_col(cols, "end time")
    c_ex1 = _find_col(cols, "exercise 1")
    c_ex2 = _find_col(cols, "exercise 2")
    c_ex3 = _find_col(cols, "exercise 3")
    c_dose = _find_col(cols, "dose / target", "dose")
    c_prog = _find_col(cols, "progression")
    c_why = _find_col(cols, "why it matters")
    c_notes = _find_col(cols, "notes")
    if None in (c_start, c_ex1, c_dose, c_why):
        raise ValueError("02_Brief_Today: required main columns missing")

    # data block ends at first 'answer to timing question' / 'non-negotiable'
    # row or first all-blank row (mirrors the reader's break conditions).
    block_end = main_hdr
    for r in range(main_hdr + 1, ws.max_row + 1):
        values = _row_text(ws, r)
        first = values[0].lower() if values else ""
        if "answer to timing question" in first or "non-negotiable" in first:
            break
        if not any(values):
            break
        block_end = r

    used_cols = [
        c for c in (c_start, c_end, c_ex1, c_ex2, c_ex3, c_dose, c_prog, c_why, c_notes)
        if c is not None
    ]
    _clear_block(ws, main_hdr + 1, block_end, used_cols)

    capacity = block_end - main_hdr
    if len(main_tasks) > capacity:
        print(
            f"[xlsx] 02_Brief_Today: {len(main_tasks)} tasks > {capacity} data "
            f"rows; writing first {capacity}, skipping overflow"
        )
    for offset, task in enumerate(main_tasks[:capacity]):
        r = main_hdr + 1 + offset
        names_raw = _meta(task, "exercise_names")
        if names_raw:
            exercises = [e.strip() for e in str(names_raw).split(";") if e.strip()]
        else:
            exercises = [e.strip() for e in str(task.get("name", "")).split("+") if e.strip()]
        timing = str(task.get("timing") or "")
        parts = re.split(r"[-–]", timing, maxsplit=1)
        start = parts[0].strip() if parts else ""
        end = parts[1].strip() if len(parts) > 1 else ""

        def put(col: int | None, value: Any) -> None:
            if col is not None and value not in (None, ""):
                ws.cell(row=r, column=col).value = value

        put(c_start, start)
        put(c_end, end)
        put(c_ex1, exercises[0] if len(exercises) > 0 else None)
        put(c_ex2, exercises[1] if len(exercises) > 1 else None)
        put(c_ex3, exercises[2] if len(exercises) > 2 else None)
        put(c_dose, task.get("target_value"))
        put(c_prog, _meta(task, "progression"))
        put(c_why, task.get("description"))
        put(c_notes, _meta(task, "notes"))

    # --- weekly minimums table (second header: pillar | minimum | week 1) ---
    min_hdr = _find_header_row(ws, ["pillar", "minimum", "week 1"], limit=ws.max_row)
    if min_hdr is None:
        return  # template has no minimums table; main table already written
    mcols = _col_map(ws, min_hdr)
    m_pillar = _find_col(mcols, "pillar")
    m_min = _find_col(mcols, "minimum")
    m_w1 = _find_col(mcols, "week 1")
    m_w2 = _find_col(mcols, "week 2")
    m_track = _find_col(mcols, "track")
    m_notes = _find_col(mcols, "notes")
    if m_pillar is None or m_min is None:
        return  # cannot safely round-trip; leave minimums as template

    m_end = min_hdr
    for r in range(min_hdr + 1, ws.max_row + 1):
        if not any(_row_text(ws, r)):
            break
        m_end = r
    m_cols = [c for c in (m_pillar, m_min, m_w1, m_w2, m_track, m_notes) if c is not None]
    _clear_block(ws, min_hdr + 1, m_end, m_cols)

    m_capacity = m_end - min_hdr
    if len(must_tasks) > m_capacity:
        print(
            f"[xlsx] 02_Brief_Today minimums: {len(must_tasks)} > {m_capacity} "
            f"rows; writing first {m_capacity}"
        )
    for offset, task in enumerate(must_tasks[:m_capacity]):
        r = min_hdr + 1 + offset
        # name is "(must) <pillar>" → recover the pillar label
        pillar_label = re.sub(r"^\(must\)\s*", "", str(task.get("name", ""))).strip()

        def putm(col: int | None, value: Any) -> None:
            if col is not None and value not in (None, ""):
                ws.cell(row=r, column=col).value = value

        putm(m_pillar, pillar_label)
        putm(m_min, task.get("target_value"))
        putm(m_w1, _meta(task, "week_1"))
        putm(m_w2, _meta(task, "week_2"))
        putm(m_track, _meta(task, "track"))
        putm(m_notes, task.get("description"))


def _write_supplements(ws: Any, tasks: list[dict]) -> None:
    """Rebuild 09_Supplements, inverting _read_supplements_from_excel."""
    supps = [
        t
        for t in tasks
        if t.get("pillar") == "supplements" and not t.get("is_reference")
    ]

    hdr = _find_header_row(ws, ["supplement", "status", "dose", "frequency"])
    if hdr is None:
        raise ValueError("09_Supplements: header not found")
    cols = _col_map(ws, hdr)

    c_name = _find_col(cols, "supplement")
    c_cat = _find_col(cols, "cat")
    c_status = _find_col(cols, "status")
    c_dose = _find_col(cols, "dose")
    c_freq = _find_col(cols, "frequency")
    c_timing = _find_col(cols, "timing")
    c_trigger = _find_col(cols, "trigger marker")
    c_skip = _find_col(cols, "skip if")
    c_stop = _find_col(cols, "stop rule")
    c_combine = _find_col(cols, "do not combine")
    c_why = _find_col(cols, "why")
    c_evidence = _find_col(cols, "evidence")
    c_source = _find_col(cols, "source key")
    if None in (c_name, c_status, c_dose, c_freq):
        raise ValueError("09_Supplements: required columns missing")

    # The reader does not STOP at a non-task-status row — it CONTINUEs past it
    # (e.g. the template's "active (if needed)" rows and the stacking-rules
    # section). To stay consistent we clear every row the reader would treat as
    # a task: scan to the last row whose status cell is a real task status.
    block_end = hdr
    for r in range(hdr + 1, ws.max_row + 1):
        status_cell = ws.cell(row=r, column=c_status).value
        status = str(status_cell).strip().lower() if status_cell is not None else ""
        name_cell = ws.cell(row=r, column=c_name).value
        if name_cell and status in _SUPPLEMENT_STATUSES:
            block_end = r

    used = [
        c
        for c in (
            c_name, c_cat, c_status, c_dose, c_freq, c_timing, c_trigger,
            c_skip, c_stop, c_combine, c_why, c_evidence, c_source,
        )
        if c is not None
    ]
    _clear_block(ws, hdr + 1, block_end, used)

    capacity = block_end - hdr
    if len(supps) > capacity:
        print(
            f"[xlsx] 09_Supplements: {len(supps)} > {capacity} data rows; "
            f"writing first {capacity}"
        )
    for offset, task in enumerate(supps[:capacity]):
        r = hdr + 1 + offset

        def put(col: int | None, value: Any) -> None:
            if col is not None and value not in (None, ""):
                ws.cell(row=r, column=col).value = value

        put(c_name, task.get("name"))
        put(c_cat, _meta(task, "category"))
        put(c_status, _meta(task, "status") or "Active")
        put(c_dose, task.get("target_value"))
        put(c_freq, task.get("schedule"))
        put(c_timing, task.get("timing"))
        put(c_trigger, _meta(task, "trigger_marker"))
        put(c_skip, _meta(task, "skip_if"))
        put(c_stop, _meta(task, "stop_rule"))
        put(c_combine, _meta(task, "do_not_combine_with"))
        put(c_why, task.get("description"))
        put(c_evidence, _meta(task, "evidence"))
        put(c_source, task.get("source_key"))


def _write_screenings(ws: Any, pillar: str, screenings: list[dict]) -> None:
    """
    Rebuild a screening sheet, inverting _read_screenings_from_excel.

    pillar is 'blood_markers' or 'screenings_safety' (different layouts).
    """
    items = [s for s in screenings if s.get("pillar") == pillar]

    if pillar == "blood_markers":
        hdr = _find_header_row(ws, ["marker", "optimal range"])
        if hdr is None:
            raise ValueError("blood_markers: header not found")
        cols = _col_map(ws, hdr)
        c_name = _find_col(cols, "marker")
        c_desc = _find_col(cols, "why it matters")
        c_target = _find_col(cols, "optimal range")
        if c_name is None:
            raise ValueError("blood_markers: marker column missing")
        used = [c for c in (c_name, c_desc, c_target) if c is not None]
    else:  # screenings_safety
        hdr = _find_header_row(ws, ["screening", "frequency"])
        if hdr is None:
            raise ValueError("screenings_safety: header not found")
        cols = _col_map(ws, hdr)
        c_name = _find_col(cols, "screening")
        c_desc = _find_col(cols, "why")
        c_target = _find_col(cols, "frequency")
        if c_name is None:
            raise ValueError("screenings_safety: screening column missing")
        used = [c for c in (c_name, c_desc, c_target) if c is not None]

    block_end = hdr
    for r in range(hdr + 1, ws.max_row + 1):
        if not any(_row_text(ws, r)):
            break
        block_end = r

    _clear_block(ws, hdr + 1, block_end, used)

    capacity = block_end - hdr
    if len(items) > capacity:
        print(
            f"[xlsx] {pillar}: {len(items)} > {capacity} data rows; "
            f"writing first {capacity}"
        )
    for offset, item in enumerate(items[:capacity]):
        r = hdr + 1 + offset

        def put(col: int | None, value: Any) -> None:
            if col is not None and value not in (None, ""):
                ws.cell(row=r, column=col).value = value

        put(c_name, item.get("name"))
        put(c_desc, item.get("description"))
        if pillar == "blood_markers":
            put(c_target, item.get("target_value"))
        else:
            months = item.get("frequency_months")
            if months is not None:
                years = round(months / 12, 2)
                # whole numbers render without a trailing .0
                put(c_target, int(years) if years == int(years) else years)


def _write_rotation(ws: Any, rotation_days: list[dict]) -> None:
    """Rebuild 05_30Day_Rotation, inverting _read_rotation_from_excel."""
    hdr = _find_header_row(ws, ["day", "week", "focus", "morning time"])
    if hdr is None:
        raise ValueError("30day_rotation: header not found")
    cols = _col_map(ws, hdr)

    c_day = _find_col(cols, "day")
    c_week = _find_col(cols, "week")
    c_block = _find_col(cols, "focus")
    c_morn = _find_col(cols, "morning time")
    c_warm = _find_col(cols, "warm-up min")
    c_upper = _find_col(cols, "upper back")
    c_secmin = _find_col(cols, "secondary min")
    c_cool = _find_col(cols, "cool-down min")
    c_total = _find_col(cols, "total min")
    c_fits = _find_col(cols, "fits 60")
    c_prio = _find_col(cols, "priority exercises")
    c_secex = _find_col(cols, "secondary exercises")
    c_rule = _find_col(cols, "week rule")
    c_notes = _find_col(cols, "notes")
    if c_day is None or c_block is None:
        raise ValueError("30day_rotation: required columns missing")

    block_end = hdr
    for r in range(hdr + 1, ws.max_row + 1):
        if not any(_row_text(ws, r)):
            break
        block_end = r

    used = [
        c
        for c in (
            c_day, c_week, c_block, c_morn, c_warm, c_upper, c_secmin, c_cool,
            c_total, c_fits, c_prio, c_secex, c_rule, c_notes,
        )
        if c is not None
    ]
    _clear_block(ws, hdr + 1, block_end, used)

    capacity = block_end - hdr
    ordered = sorted(rotation_days, key=lambda d: d.get("day_number") or 0)
    if len(ordered) > capacity:
        print(
            f"[xlsx] 05_30Day_Rotation: {len(ordered)} > {capacity} data rows; "
            f"writing first {capacity}"
        )
    for offset, day in enumerate(ordered[:capacity]):
        r = hdr + 1 + offset

        def put(col: int | None, value: Any) -> None:
            if col is not None and value not in (None, ""):
                ws.cell(row=r, column=col).value = value

        put(c_day, day.get("day_number"))
        put(c_week, day.get("week_number"))
        put(c_block, day.get("block_name"))
        put(c_morn, day.get("morning_time"))
        put(c_warm, day.get("warm_up_min"))
        put(c_upper, day.get("upper_back_core_min"))
        put(c_secmin, day.get("secondary_min"))
        put(c_cool, day.get("cool_down_min"))
        put(c_total, day.get("total_min"))
        put(c_fits, day.get("fits_60"))
        put(c_prio, day.get("priority_exercises"))
        put(c_secex, day.get("secondary_exercises"))
        put(c_rule, day.get("week_rule"))
        put(c_notes, day.get("notes"))


def generate_xlsx(workbook_json: dict) -> bytes:
    """
    Load the template xlsx, rebuild the editable sheets from ``workbook_json``,
    and return raw bytes.

    Stamps ``01_Personal_Settings`` then repopulates the brief-today,
    supplements, blood-markers, screenings-safety, and 30-day rotation sheets
    from the canonical JSON arrays. Each sheet's rewrite is independently
    fault-isolated: a missing sheet or unexpected layout skips that one sheet
    (logged) and leaves it as the template — the download never 500s.

    Raises FileNotFoundError only if the template itself is missing.
    """
    template = get_or_copy_template()
    wb = openpyxl.load_workbook(template)
    profile = workbook_json.get("personal_settings", {}) or {}
    tasks = workbook_json.get("tasks", []) or []
    rotation_days = workbook_json.get("rotation_days", []) or []
    screenings = workbook_json.get("screenings", []) or []

    # Populate 01_Personal_Settings sheet if present
    if "01_Personal_Settings" in wb.sheetnames:
        ws = wb["01_Personal_Settings"]
        # Write known cell positions; fall back silently if structure differs
        try:
            ws["B6"] = profile.get("name", "")
            ws["B7"] = profile.get("date_of_birth", "")
            ws["B8"] = profile.get("sex", "")
            ws["B9"] = profile.get("height_cm", "")
            ws["B10"] = profile.get("weight_kg", "")
        except Exception:
            # Non-fatal: template structure may differ from expected layout
            pass

    # Each sheet rewrite is independently fault-isolated (§11): one bad sheet
    # must not abort the others or the download.
    rewrites: list[tuple[str, str, Any]] = [
        ("brief_today", "02_Brief_Today", lambda w: _write_brief_today(w, tasks)),
        ("supplements", "09_Supplements", lambda w: _write_supplements(w, tasks)),
        (
            "blood_markers",
            "10_Blood_Markers",
            lambda w: _write_screenings(w, "blood_markers", screenings),
        ),
        (
            "screenings_safety",
            "11_Screenings_Safety",
            lambda w: _write_screenings(w, "screenings_safety", screenings),
        ),
        ("30day_rotation", "05_30Day_Rotation", lambda w: _write_rotation(w, rotation_days)),
    ]
    for pillar, label, writer in rewrites:
        try:
            ws = _find_sheet(wb, pillar)
            if ws is None:
                print(f"[xlsx] sheet for '{pillar}' ({label}) not found; left as template")
                continue
            writer(ws)
        except Exception as e:  # noqa: BLE001 — resilience by design (§11)
            print(f"[xlsx] skipping '{label}' rewrite (shape mismatch): {e}")

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()
