"""
AI-powered plan ingestion service.

Flow: upload .xlsx or .json
      → extract readable text from every sheet
      → call OpenAI with structured-output schema
      → save Plan + TaskTemplates + RotationDays + Screenings + plan_json to DB
      → enrich brief_today blocks with per-exercise details from exercise_library (Python, no LLM)
      → pre-generate DailyTodo rows for the next 30 days
      → return IngestResult summary

The LLM handles intent understanding; Python handles structured cross-sheet merging.
"""
import io
import json
import os
import re
import urllib.parse
import uuid
from datetime import date, timedelta
from typing import TYPE_CHECKING, BinaryIO

if TYPE_CHECKING:
    from langchain_openai import ChatOpenAI

import openpyxl
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from app.constants import DEFAULT_USER_ID, SKIP_SHEETS
from app.db.models import DailyTodo, Plan, RotationDay, Screening, TaskTemplate
from app.services.scheduler import _is_scheduled_today


# ── Structured output schema the LLM must return ─────────────────────────

class IngestedTask(BaseModel):
    pillar: str = Field(
        description=(
            "Sheet name normalized: lowercase, spaces→underscores, "
            "strip leading number+underscore prefix. "
            "E.g. '02_Brief_Today' → 'brief_today', '09_Supplements' → 'supplements'."
        )
    )
    name: str = Field(description="Task or item name.")
    description: str | None = Field(None, description="Why this task matters.")
    schedule: str | None = Field(
        None,
        description=(
            "Normalized schedule. Use one of: 'daily', '2x/week', '3x/week', "
            "'Mon/Wed/Fri', 'Mon/Thu', 'weekly', 'as needed'. "
            "Default to 'daily' for brief_today and supplement items with no explicit schedule."
        ),
    )
    timing: str | None = Field(None, description="Time of day: 'Morning', 'Evening', '5:40 AM', etc.")
    target_value: str | None = Field(
        None,
        description=(
            "Target amount or dosage. For brief_today blocks: copy the FULL text from "
            "'Week 1 easy start' column verbatim — this is the exercise list (e.g. "
            "'2 rounds: band row, wall slide, dead bug/bird dog')."
        ),
    )
    unit: str | None = Field(None, description="Unit for target_value.")
    how_to: str | None = Field(None, description="Step-by-step execution instructions.")
    why_mechanism: str | None = Field(None, description="Scientific rationale.")
    safety_notes: str | None = Field(None, description="Safety warnings or pain-stop rules.")
    video_link: str | None = Field(
        None,
        description="YouTube URL only if an actual URL is present. Leave null for search terms.",
    )
    link: str | None = Field(
        None,
        description="GIF/reference URL only if an actual URL is present. Leave null for search terms.",
    )
    benefit_tags: str | None = Field(None, description="Health benefits, comma-separated.")
    source_key: str | None = Field(None, description="Source reference key.")
    extra_metadata: dict[str, str] | None = Field(
        None,
        description="Structured catch-all fields copied from deterministic Excel readers.",
    )
    is_reference: bool = Field(
        False,
        description=(
            "Set True ONLY for: nutrition, sleep_recovery, cognitive_social, exercise_library. "
            "Everything else is False — brief_today, supplements, etc. are actionable daily tasks."
        ),
    )


class IngestedRotationDay(BaseModel):
    day_number: int = Field(description="Day number 1–30.")
    block_name: str = Field(description="Workout session name or focus.")
    week_number: int | None = None
    morning_time: str | None = Field(None, description="Session start time e.g. '5:40 AM'.")
    warm_up_min: str | None = Field(None, description="Warm-up minutes e.g. '8'.")
    upper_back_core_min: str | None = Field(None, description="Upper back+core block minutes.")
    secondary_min: str | None = Field(None, description="Secondary block minutes.")
    cool_down_min: str | None = Field(None, description="Cool-down minutes.")
    total_min: str | None = Field(None, description="Total session minutes.")
    fits_60: str | None = Field(None, description="'Yes' if session fits 60-min window.")
    priority_exercises: str | None = Field(None, description="Comma-separated priority exercise names.")
    secondary_exercises: str | None = Field(None, description="Comma-separated secondary exercise names.")
    week_rule: str | None = Field(None, description="Intensity rule e.g. 'Week 1: Easy / RPE 5-6'.")
    notes: str | None = None
    # v3 fields — populated when workbook uses the older layout
    warm_up: str | None = None
    priority_block: str | None = None
    secondary_block: str | None = None
    cardio_steps: str | None = None
    cool_down: str | None = None
    nutrition_focus: str | None = None
    intensity_cap: str | None = None


class IngestedScreening(BaseModel):
    pillar: str = Field(description="'blood_markers' or 'screenings_safety'.")
    name: str
    description: str | None = None
    frequency_months: int | None = Field(None, description="12=annual, 6=biannual, 3=quarterly.")
    target_value: str | None = Field(None, description="Optimal range for blood markers.")


class IngestedPlan(BaseModel):
    plan_name: str = Field(description="A short descriptive name for this plan.")
    tasks: list[IngestedTask]
    # rotation_days and screenings are extracted directly from Excel — not via LLM


# ── Screening / rotation sheet names (normalized) ─────────────────────────

_SCREENING_SHEET_NAMES: frozenset[str] = frozenset({'blood_markers', 'screenings_safety'})
_ROTATION_SHEET_NAME = '30day_rotation'


# ── File extraction ───────────────────────────────────────────────────────

def _normalize_pillar(sheet_name: str) -> str:
    name = re.sub(r'^\d+_?', '', sheet_name.strip())
    return name.lower().replace(' ', '_')


def _ws_to_text(ws, sheet_name: str, pillar: str) -> str:
    """Convert a single openpyxl worksheet to an LLM-readable text block."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return ''

    best_idx, best_count = 0, 0
    for i, row in enumerate(rows[:15]):
        count = sum(1 for c in row if c is not None)
        if count > best_count:
            best_count, best_idx = count, i

    raw_header = rows[best_idx]
    header = [str(c).strip() if c is not None else '' for c in raw_header]
    while header and not header[-1]:
        header.pop()
    col_count = len(header)

    lines = [
        f"=== SHEET: {sheet_name} (pillar: {pillar}) ===",
        ' | '.join(header),
        '-' * 60,
    ]

    for row in rows[best_idx + 1:]:
        if not any(c is not None for c in row):
            continue
        cells = [str(c).strip() if c is not None else '' for c in row[:col_count]]
        if not any(cells):
            continue
        lines.append(' | '.join(cells))

    return '\n'.join(lines)


def _extract_sheet_texts(file_content: bytes) -> dict[str, str]:
    """Return {pillar: text} for each non-skipped task sheet (one entry per sheet)."""
    wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    result: dict[str, str] = {}
    for sheet_name in wb.sheetnames:
        pillar = _normalize_pillar(sheet_name)
        if pillar in SKIP_SHEETS:
            continue
        text = _ws_to_text(wb[sheet_name], sheet_name, pillar)
        if text:
            result[pillar] = text
    return result


def extract_xlsx_text(file: BinaryIO) -> str:
    """Concatenate all non-skip sheets into one text block (kept for JSON / legacy path)."""
    file_content = file.read()
    return '\n\n'.join(_extract_sheet_texts(file_content).values())


def extract_json_text(file: BinaryIO) -> str:
    """Read a JSON file and return it as a formatted string."""
    raw = file.read()
    if isinstance(raw, bytes):
        raw = raw.decode('utf-8')
    try:
        return json.dumps(json.loads(raw), indent=2)
    except Exception:
        return str(raw)


# ── Exercise library direct reader (no LLM) ───────────────────────────────

# Matches =HYPERLINK("url", ...) formula so we can extract the URL when data_only caches
# the display text instead of the formula string.
_HYPERLINK_FORMULA_RE = re.compile(r'=HYPERLINK\s*\(\s*"([^"]+)"', re.I)


def _to_search_url(value: str | None, base_url: str) -> str | None:
    """Return value unchanged if it is already a URL; otherwise build a search URL from the text."""
    if not value:
        return None
    v = value.strip()
    if not v:
        return None
    if v.lower().startswith('http'):
        return v
    return base_url + urllib.parse.quote_plus(v)


def _read_exercise_library(file_content: bytes) -> dict[str, dict]:
    """
    Read the exercise_library sheet directly from the Excel file.
    Returns a lookup dict keyed by lowercase exercise name with full detail fields.

    Loaded without read_only so that cell.hyperlink is accessible — the YouTube and
    GIF columns store the actual URL as an Excel hyperlink whose display text is just
    the column header (e.g. "YouTube Demo Search").  read_only=True silently drops all
    hyperlink metadata, leaving only the display text.
    """
    # data_only=False keeps formula strings (e.g. =HYPERLINK("url","text")) so the
    # regex in _cell_link can extract the real URL. data_only=True would replace the
    # formula with the cached display text ("YouTube demo search"), losing the URL.
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=False)
    library: dict[str, dict] = {}

    for sheet_name in wb.sheetnames:
        if 'exercise' not in sheet_name.lower() or 'library' not in sheet_name.lower():
            continue

        ws = wb[sheet_name]
        # Keep cell objects (not values_only) so we can read .hyperlink
        rows = list(ws.iter_rows(values_only=False))

        # Find header row (densest row in first 15)
        best_idx, best_count = 0, 0
        for i, row in enumerate(rows[:15]):
            count = sum(1 for c in row if c.value is not None)
            if count > best_count:
                best_count, best_idx = count, i

        header = [str(c.value).strip().lower() if c.value is not None else '' for c in rows[best_idx]]

        def col_idx(keyword: str) -> int | None:
            for i, h in enumerate(header):
                if keyword.lower() in h:
                    return i
            return None

        idx = {
            'name':             col_idx('exercise'),
            'category':         col_idx('category'),
            'setup':            col_idx('setup'),
            'starting_position':col_idx('starting position'),
            'how_to':           col_idx('step-by-step'),
            'bracing_cue':      col_idx('core/bracing'),
            'common_mistakes':  col_idx('common mistakes'),
            'week1_dosage':     col_idx('week 1 dosage'),
            'safety_notes':     col_idx('safety stop'),
            'why_it_matters':   col_idx('why it matters'),
            'video_link':       col_idx('youtube'),
            'gif_link':         col_idx('animated gif'),
        }

        if idx['name'] is None:
            break

        def _cell(row: tuple, key: str) -> str | None:
            i = idx.get(key)
            if i is None or i >= len(row):
                return None
            cell = row[i]
            if cell.value is None:
                return None
            v = str(cell.value).strip()
            return v if v else None

        def _cell_link(row: tuple, key: str) -> str | None:
            """Read a link column: hyperlink target → HYPERLINK formula → raw cell value."""
            i = idx.get(key)
            if i is None or i >= len(row):
                return None
            cell = row[i]
            # 1. Actual cell hyperlink (Insert Hyperlink / openpyxl relationship)
            if cell.hyperlink and cell.hyperlink.target:
                v = cell.hyperlink.target.strip()
                return v or None
            # 2. =HYPERLINK("url", "text") formula (data_only gives cached display text,
            #    but without data_only we'd see the formula — handle both)
            if isinstance(cell.value, str):
                m = _HYPERLINK_FORMULA_RE.match(cell.value.strip())
                if m:
                    v = m.group(1).strip()
                    return v or None
            # 3. Plain cell value — could be a URL or a search-term string
            if cell.value is None:
                return None
            v = str(cell.value).strip()
            return v if v else None

        for row in rows[best_idx + 1:]:
            name_i = idx['name']
            if not any(c.value is not None for c in row):
                continue
            if name_i is None or name_i >= len(row) or row[name_i].value is None:
                continue
            name = str(row[name_i].value).strip()
            if not name:
                continue
            library[name.lower()] = {
                'name':             name,
                'category':         _cell(row, 'category'),
                'setup':            _cell(row, 'setup'),
                'starting_position':_cell(row, 'starting_position'),
                'how_to':           _cell(row, 'how_to'),
                'bracing_cue':      _cell(row, 'bracing_cue'),
                'common_mistakes':  _cell(row, 'common_mistakes'),
                'week1_dosage':     _cell(row, 'week1_dosage'),
                'safety_notes':     _cell(row, 'safety_notes'),
                'why_it_matters':   _cell(row, 'why_it_matters'),
                'video_link': _to_search_url(
                    _cell_link(row, 'video_link'),
                    'https://www.youtube.com/results?search_query=',
                ),
                'gif_link': _to_search_url(
                    _cell_link(row, 'gif_link'),
                    'https://giphy.com/search/',
                ),
            }
        break  # only process first matching sheet

    return library


def _clean_cell_text(value: object | None) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return re.sub(r'^[•*\-\s]+', '', text).strip() or None


def _read_brief_today_from_excel(file_content: bytes) -> list[IngestedTask]:
    """Parse the v5 02_Brief_Today two-table layout directly from Excel."""
    wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)

    ws = None
    for sheet_name in wb.sheetnames:
        if _normalize_pillar(sheet_name) == 'brief_today':
            ws = wb[sheet_name]
            break
    if ws is None:
        return []

    rows = list(ws.iter_rows(values_only=True))
    tasks: list[IngestedTask] = []

    def row_text(row: tuple) -> list[str]:
        return [str(c).strip() if c is not None else '' for c in row]

    def find_header(required: list[str], start_at: int = 0) -> int | None:
        for i, row in enumerate(rows[start_at:], start=start_at):
            lowered = [c.lower() for c in row_text(row)]
            if all(any(req in cell for cell in lowered) for req in required):
                return i
        return None

    def col_map(header_idx: int) -> dict[str, int]:
        header = row_text(rows[header_idx])
        return {h.lower(): i for i, h in enumerate(header) if h}

    def cell(row: tuple, columns: dict[str, int], name: str) -> str | None:
        idx = columns.get(name.lower())
        if idx is None or idx >= len(row):
            return None
        return _clean_cell_text(row[idx])

    main_header_idx = find_header(['start time', 'exercise 1', 'dose'])
    if main_header_idx is not None:
        columns = col_map(main_header_idx)
        for row in rows[main_header_idx + 1:]:
            values = row_text(row)
            first = values[0].lower() if values else ''
            if not any(values):
                continue
            if 'answer to timing question' in first or 'non-negotiable' in first:
                break

            start = cell(row, columns, 'start time')
            end = cell(row, columns, 'end time')
            dose = cell(row, columns, 'dose / target')
            exercises = [
                ex for ex in (
                    cell(row, columns, 'exercise 1'),
                    cell(row, columns, 'exercise 2'),
                    cell(row, columns, 'exercise 3'),
                )
                if ex
            ]
            if not (start or end or dose or exercises):
                continue

            name = ' + '.join(exercises) if exercises else 'Training log'
            meta: dict[str, str] = {}
            if exercises:
                meta['exercise_names'] = '; '.join(exercises)
            progression = cell(row, columns, 'progression')
            notes = cell(row, columns, 'notes')
            if progression:
                meta['progression'] = progression
            if notes:
                meta['notes'] = notes

            tasks.append(IngestedTask(
                pillar='brief_today',
                name=name,
                description=cell(row, columns, 'why it matters'),
                schedule='daily',
                timing='-'.join(part for part in (start, end) if part),
                target_value=dose,
                extra_metadata=meta or None,
                is_reference=False,
            ))

    minimums_header_idx = find_header(['pillar', 'minimum', 'week 1'])
    if minimums_header_idx is not None:
        columns = col_map(minimums_header_idx)
        for row in rows[minimums_header_idx + 1:]:
            pillar = cell(row, columns, 'pillar')
            minimum = cell(row, columns, 'minimum')
            if not pillar or not minimum:
                continue
            meta: dict[str, str] = {'must': 'true'}
            for source, target in (
                ('week 1', 'week_1'),
                ('week 2', 'week_2'),
                ('track', 'track'),
            ):
                value = cell(row, columns, source)
                if value:
                    meta[target] = value
            notes = cell(row, columns, 'notes')

            tasks.append(IngestedTask(
                pillar='brief_today',
                name=f'(must) {pillar}',
                description=notes,
                schedule='weekly',
                target_value=minimum,
                how_to=meta.get('week_1'),
                extra_metadata=meta,
                is_reference=False,
            ))

    return tasks


def _normalize_supplement_schedule(frequency: str | None) -> str:
    """Map supplements frequency text into scheduler-friendly values."""
    if not frequency:
        return "daily"
    f = frequency.strip().lower()
    if "daily" in f:
        return "daily"
    if "alternate day" in f or "eod" in f or "every other day" in f:
        return "every other day"
    if "2×/wk" in f or "2x/week" in f or "2x/wk" in f:
        return "2x/week"
    if "only if prescribed" in f or "specialist" in f:
        return "only if prescribed"
    if "as needed" in f:
        return "as needed"
    if "cycle" in f or "on / off" in f or "on/off" in f:
        return "cycle"
    return frequency.strip()


def _read_supplements_from_excel(file_content: bytes) -> list[IngestedTask]:
    """Parse 09_Supplements directly so status/category and stop rules are deterministic."""
    wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)

    ws = None
    for sheet_name in wb.sheetnames:
        if _normalize_pillar(sheet_name) == 'supplements':
            ws = wb[sheet_name]
            break
    if ws is None:
        return []

    rows = list(ws.iter_rows(values_only=True))
    tasks: list[IngestedTask] = []

    def row_text(row: tuple) -> list[str]:
        return [str(c).strip() if c is not None else '' for c in row]

    def find_header(required: list[str]) -> int | None:
        for i, row in enumerate(rows[:20]):
            lowered = [c.lower() for c in row_text(row)]
            if all(any(req in cell for cell in lowered) for req in required):
                return i
        return None

    header_idx = find_header(['supplement', 'status', 'dose', 'frequency'])
    if header_idx is None:
        return []

    header = [h.lower() for h in row_text(rows[header_idx])]

    def _col(keyword: str) -> int | None:
        for i, h in enumerate(header):
            if keyword in h:
                return i
        return None

    col_name = _col('supplement')
    col_cat = _col('cat')
    col_status = _col('status')
    col_dose = _col('dose')
    col_frequency = _col('frequency')
    col_timing = _col('timing')
    col_trigger = _col('trigger marker')
    col_skip = _col('skip if')
    col_stop = _col('stop rule')
    col_combine = _col('do not combine')
    col_why = _col('why')
    col_evidence = _col('evidence')
    col_source = _col('source key')

    def _cell(row: tuple, col: int | None) -> str | None:
        if col is None or col >= len(row):
            return None
        return _clean_cell_text(row[col])

    for row in rows[header_idx + 1:]:
        name = _cell(row, col_name)
        status = _cell(row, col_status)
        if not name or not status:
            continue

        # stop at other sections (stacking rules / algorithm) where status column is not task status
        if status.lower() not in {'active', 'discuss', 'caution', 'avoid', 'deprioritize', 'optional', 'specialist only'}:
            continue

        category = _cell(row, col_cat)
        dose = _cell(row, col_dose)
        frequency = _cell(row, col_frequency)
        timing = _cell(row, col_timing)
        trigger = _cell(row, col_trigger)
        skip_if = _cell(row, col_skip)
        stop_rule = _cell(row, col_stop)
        do_not_combine = _cell(row, col_combine)
        why = _cell(row, col_why)
        evidence = _cell(row, col_evidence)
        source_key = _cell(row, col_source)

        meta: dict[str, str] = {'status': status}
        if category:
            meta['category'] = category
        if trigger:
            meta['trigger_marker'] = trigger
        if skip_if:
            meta['skip_if'] = skip_if
        if stop_rule:
            meta['stop_rule'] = stop_rule
        if do_not_combine:
            meta['do_not_combine_with'] = do_not_combine
        if evidence:
            meta['evidence'] = evidence

        how_to_parts = [p for p in (skip_if, stop_rule, do_not_combine) if p]

        tasks.append(IngestedTask(
            pillar='supplements',
            name=name,
            description=why,
            schedule=_normalize_supplement_schedule(frequency),
            timing=timing,
            target_value=dose,
            how_to='\n'.join(how_to_parts) if how_to_parts else None,
            source_key=source_key,
            extra_metadata=meta,
            is_reference=False,
        ))

    return tasks


def _find_exercise(name: str, library: dict[str, dict]) -> dict | None:
    """Fuzzy match an exercise name against the library lookup dict."""
    name_lower = name.strip().lower()
    # Exact
    if name_lower in library:
        return library[name_lower]
    # Substring both ways
    for k, v in library.items():
        if name_lower in k or k in name_lower:
            return v
    # Any significant word (>4 chars) appears in a library key
    words = [w for w in re.split(r'\W+', name_lower) if len(w) > 4]
    for word in words:
        for k, v in library.items():
            if word in k:
                return v
    return None


# Maps keywords in block names to exercise_library Category values
_CATEGORY_KEYWORDS: dict[str, str] = {
    'lower body': 'Lower body',
    'push':       'Push',
    'carry':      'Core/Carry',
    'zone 2':     'Cardio',
    'cardio':     'Cardio',
    'balance':    'Balance',
    'cool-down':  'Cool-down',
    'warm-up':    'Warm-up',
}

# Regex that flags a text chunk as a dosage/instruction rather than an exercise name
_INSTRUCTION_RE = re.compile(
    r'\bRPE\b|technique\s|move\s+slow|easy\s+sets|no\s+aggress|long\s+exhale'
    r'|quick\s+note|log\s+pain|\bmin\b.*\bRPE\b',
    re.I,
)

# Regex that flags a "/" sub-part as a measurement unit (don't split "sec/side")
_UNIT_RE = re.compile(r'\b(sec|min|rep|set|side|kg|lb|ml|oz)\b|\d', re.I)


def _find_exercises_for_block(block: TaskTemplate, library: dict[str, dict]) -> list[dict]:
    """
    Find exercises for a brief_today workout block using four sources in order.
    Only returns exercises that actually match the library — no phantom names stored.

    Sources:
      1. target_value  "Week 1 easy start": "2 rounds: band row, wall slide, dead bug"
      2. how_to        only when it has an explicit list: "like sit-to-stand, glute bridge"
      3. task name     "+" notation: "easy nasal walk + shoulder circles + cat-cow"
      4. category      keywords in task name mapped to exercise_library Category column
    """
    seen: set[str] = set()
    exercises: list[dict] = []

    def _add_candidate(candidate: str) -> None:
        candidate = _clean_cell_text(candidate) or ''
        # Too short or pure dosage like "1×20–30" or "sec" — skip
        if len(candidate) <= 3:
            return
        if re.match(r'^[\d×xX\-–\s]+$', candidate):
            return
        data = _find_exercise(candidate, library)
        if data and data['name'] not in seen:
            seen.add(data['name'])
            exercises.append(data)

    # Source 0 — v5 brief_today Exercise 1/2/3 columns captured in metadata.
    if block.extra_metadata:
        try:
            meta = json.loads(block.extra_metadata)
        except ValueError:
            meta = {}
        if isinstance(meta, dict):
            explicit = meta.get('exercise_names')
            if isinstance(explicit, str):
                for part in re.split(r';|,|\band\b|\bor\b', explicit):
                    _add_candidate(part)
            for key, value in meta.items():
                if str(key).lower().startswith('exercise') and isinstance(value, str):
                    _add_candidate(value)

    # Source 1 — target_value (legacy Week 1 easy start column)
    # Split on comma/and/or. Also split "/" when both sides look like exercise names
    # (e.g. "dead bug/bird dog") but NOT when "/" is a unit separator ("sec/side").
    if block.target_value:
        text = re.sub(r'^\d+\s+rounds?:\s*', '', block.target_value, flags=re.I)
        if not _INSTRUCTION_RE.search(text):
            for chunk in re.split(r',|\band\b|\bor\b', text):
                if '/' in chunk:
                    parts = [p.strip() for p in chunk.split('/')]
                    if not any(_UNIT_RE.search(p) for p in parts):
                        for p in parts:
                            _add_candidate(p)
                        continue
                _add_candidate(chunk)

    # Source 2 — how_to: ONLY scan when it contains an explicit example list
    # ("like X, Y", "e.g. X, Y", "such as X, Y") — avoids matching generic instructions
    if block.how_to:
        m = re.search(
            r'(?:like|e\.?g\.?|such as)\s+(.+?)(?:[.;]|$)',
            block.how_to, re.I,
        )
        if m:
            for part in re.split(r',|\band\b|\bor\b', m.group(1)):
                _add_candidate(part)

    # Source 3 — "+" notation in task name; split on "+" individually
    # Strip prefix up to colon only — preserve "easy walk" as a full exercise name
    if '+' in block.name:
        text = re.sub(r'^[^:]+:\s*', '', block.name, flags=re.I)
        for part in text.split('+'):
            _add_candidate(part)

    # Source 4 — category keywords in task name (fallback when no explicit exercise names found)
    # Useful for "Secondary block: lower body, push, carry, or Zone 2 depending on day"
    if not exercises:
        name_lower = block.name.lower()
        for keyword, category in _CATEGORY_KEYWORDS.items():
            if keyword in name_lower:
                for ex in library.values():
                    if ex.get('category') == category and ex['name'] not in seen:
                        seen.add(ex['name'])
                        exercises.append(ex)

    return exercises


def _enrich_block_exercises(
    db: Session, plan_id: str, library: dict[str, dict]
) -> None:
    """
    Post-processing step: cross-reference brief_today workout blocks against the
    exercise_library (read directly from Excel) to build per-exercise details.
    Stores result as JSON in TaskTemplate.exercises_json.
    """
    if not library:
        return

    blocks = (
        db.query(TaskTemplate)
        .filter(TaskTemplate.plan_id == plan_id, TaskTemplate.pillar == 'brief_today')
        .all()
    )

    for block in blocks:
        exercises = _find_exercises_for_block(block, library)

        if exercises:
            block.exercises_json = json.dumps(exercises)

    db.flush()


# ── Python-direct readers for screening and rotation sheets ──────────────


def _read_screenings_from_excel(file_content: bytes) -> list[IngestedScreening]:
    """Parse blood_markers and screenings_safety directly from Excel — no LLM needed."""
    wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    screenings: list[IngestedScreening] = []

    for sheet_name in wb.sheetnames:
        pillar = _normalize_pillar(sheet_name)
        if pillar not in _SCREENING_SHEET_NAMES:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Find widest header row in first 15 rows
        best_idx, best_count = 0, 0
        for i, row in enumerate(rows[:15]):
            count = sum(1 for c in row if c is not None)
            if count > best_count:
                best_count, best_idx = count, i

        header = [str(c).strip().lower() if c is not None else '' for c in rows[best_idx]]

        def _col(keywords: list[str]) -> int | None:
            for kw in keywords:
                for i, h in enumerate(header):
                    if kw.lower() in h:
                        return i
            return None

        def _cell(row: tuple, col: int | None) -> str | None:
            if col is None or col >= len(row) or row[col] is None:
                return None
            v = str(row[col]).strip()
            return v if v else None

        if pillar == 'blood_markers':
            name_col   = _col(['marker'])
            desc_col   = _col(['why it matters'])
            target_col = _col(['optimal range'])
            freq_months_default = 12  # annual blood work
            for row in rows[best_idx + 1:]:
                if not any(c is not None for c in row):
                    continue
                name = _cell(row, name_col)
                if not name or name.lower() in ('marker', 'panel'):
                    continue
                screenings.append(IngestedScreening(
                    pillar=pillar,
                    name=name,
                    description=_cell(row, desc_col),
                    frequency_months=freq_months_default,
                    target_value=_cell(row, target_col),
                ))
        else:  # screenings_safety
            name_col  = _col(['screening'])
            desc_col  = _col(['why'])
            freq_col  = _col(['frequency'])
            for row in rows[best_idx + 1:]:
                if not any(c is not None for c in row):
                    continue
                name = _cell(row, name_col)
                if not name:
                    continue
                # Frequency column is in years — convert to months
                freq_months: int | None = None
                freq_raw = _cell(row, freq_col)
                if freq_raw:
                    m = re.search(r'(\d+(?:\.\d+)?)', freq_raw)
                    if m:
                        freq_months = round(float(m.group(1)) * 12)
                screenings.append(IngestedScreening(
                    pillar=pillar,
                    name=name,
                    description=_cell(row, desc_col),
                    frequency_months=freq_months,
                    target_value=None,
                ))

    return screenings


def _read_rotation_from_excel(file_content: bytes) -> list[IngestedRotationDay]:
    """Parse the 05_30Day_Rotation sheet directly from Excel — no LLM needed."""
    wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        if _normalize_pillar(sheet_name) != _ROTATION_SHEET_NAME:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        best_idx, best_count = 0, 0
        for i, row in enumerate(rows[:15]):
            count = sum(1 for c in row if c is not None)
            if count > best_count:
                best_count, best_idx = count, i

        header = [str(c).strip().lower() if c is not None else '' for c in rows[best_idx]]

        def _col(keywords: list[str]) -> int | None:
            for kw in keywords:
                for i, h in enumerate(header):
                    if kw.lower() in h:
                        return i
            return None

        def _cell(row: tuple, col: int | None) -> str | None:
            if col is None or col >= len(row) or row[col] is None:
                return None
            v = str(row[col]).strip()
            return v if v else None

        def _int_cell(row: tuple, col: int | None) -> int | None:
            v = _cell(row, col)
            if v is None:
                return None
            try:
                return int(float(v))
            except (ValueError, TypeError):
                return None

        # Exact column names found in the real sheet:
        # Day | Week | Focus | Morning time | Warm-up min | Upper back + core min |
        # Secondary min | Cool-down min | Total min | Fits 60? |
        # Priority exercises | Secondary exercises | Week rule | Notes
        day_col    = _col(['day'])
        week_col   = _col(['week'])
        block_col  = _col(['focus'])
        morn_col   = _col(['morning time'])
        warmup_col = _col(['warm-up min'])
        upper_col  = _col(['upper back'])
        sec_min_col = _col(['secondary min'])
        cool_col   = _col(['cool-down min'])
        total_col  = _col(['total min'])
        fits60_col = _col(['fits 60'])
        prio_ex_col = _col(['priority exercises'])
        sec_ex_col  = _col(['secondary exercises'])
        weekrule_col = _col(['week rule'])
        notes_col  = _col(['notes'])

        days: list[IngestedRotationDay] = []
        for row in rows[best_idx + 1:]:
            if not any(c is not None for c in row):
                continue
            day_num = _int_cell(row, day_col)
            if day_num is None:
                continue
            days.append(IngestedRotationDay(
                day_number=day_num,
                week_number=_int_cell(row, week_col),
                block_name=_cell(row, block_col) or f'Day {day_num}',
                morning_time=_cell(row, morn_col),
                warm_up_min=_cell(row, warmup_col),
                upper_back_core_min=_cell(row, upper_col),
                secondary_min=_cell(row, sec_min_col),
                cool_down_min=_cell(row, cool_col),
                total_min=_cell(row, total_col),
                fits_60=_cell(row, fits60_col),
                priority_exercises=_cell(row, prio_ex_col),
                secondary_exercises=_cell(row, sec_ex_col),
                week_rule=_cell(row, weekrule_col),
                notes=_cell(row, notes_col),
            ))
        return days

    return []


# ── Per-sheet LLM extraction (primary path for xlsx) ─────────────────────

class _SheetTasksOut(BaseModel):
    tasks: list[IngestedTask]


_SHEET_TASK_PROMPT = """\
Extract EVERY non-header, non-empty row from this single Excel sheet as IngestedTask objects.

FIELD RULES
───────────
name        Short clean name. Strip time prefix: "6:30 AM — Breathing" → "Breathing".
timing      Clock time or label from a time/when column or the task name prefix.
target_value  For brief_today blocks: copy "Week 1 easy start" column VERBATIM (the exercise list).
              For supplements: dosage. For exercise_library: reps/dosage.
description The WHY — one benefit per line. Use "Why" or "Benefits" column content.
how_to      Execution steps. For workout blocks: block-level guidance only.
safety_notes Pain rules, contraindications. For brief_today: "If behind schedule" content.
schedule    Normalize to: daily | 2x/week | 3x/week | Mon/Wed/Fri | Mon/Thu | weekly | as needed.
            Default to 'daily' for brief_today and supplement rows with no explicit schedule.
is_reference True ONLY for: nutrition, sleep_recovery, cognitive_social, exercise_library.
             False for brief_today, supplements, and everything else.
video_link / link  Only real URLs starting with http. Leave null for search terms.
pillar      Set to the pillar value shown in the sheet header line — do not change it.

Extract EVERY row. Do not skip any row. Do not invent data not present in the sheet.
"""


def _extract_sheet_tasks(llm: "ChatOpenAI", sheet_text: str, pillar: str) -> list[IngestedTask]:
    """Call the LLM to extract tasks from a single sheet. Forces correct pillar."""
    from langchain_core.messages import HumanMessage, SystemMessage

    structured = llm.with_structured_output(_SheetTasksOut)
    try:
        result = structured.invoke([
            SystemMessage(content=_SHEET_TASK_PROMPT),
            HumanMessage(content=sheet_text),
        ])
        tasks = result.tasks if result and result.tasks else []
    except Exception as e:
        print(f"[ingest] error extracting {pillar}: {e}")
        tasks = []

    # Force pillar — do not trust the LLM to set it correctly
    for t in tasks:
        t.pillar = pillar

    return tasks


def _make_llm(max_tokens: int = 8000):
    """Create LLM client inferred from model name: gemini-* → Google, else → OpenAI.

    Provider is detected from GEMINI_LLM_INGEST_MODEL.
    For Google: GEMINI_LLM_MODEL overrides the specific model (falls back to GEMINI_LLM_INGEST_MODEL).
    API key: checks GOOGLE_API_KEY then GEMINI_API_KEY.
    """
    ingest_model = os.getenv("GEMINI_LLM_INGEST_MODEL", "gemini-2.5-flash-lite")
    if ingest_model.startswith("gemini"):
        from langchain_google_genai import ChatGoogleGenerativeAI
        gemini_model = os.getenv("GEMINI_LLM_MODEL", ingest_model)
        api_key = os.getenv("GOOGLE_API_KEY") or os.getenv("GEMINI_API_KEY")
        print(f"[ingest] provider=google model={gemini_model}")
        return ChatGoogleGenerativeAI(
            model=gemini_model,
            google_api_key=api_key,
            max_output_tokens=max_tokens,
            temperature=0,
        )
    else:
        from langchain_openai import ChatOpenAI
        print(f"[ingest] provider=openai model={ingest_model}")
        return ChatOpenAI(model=ingest_model, max_tokens=max_tokens, temperature=0)


# ── LLM normalization (kept for JSON / legacy path) ───────────────────────

_SYSTEM_PROMPT = """\
You are a data extraction assistant for a personal longevity and fitness plan. \
Parse the content and return a structured IngestedPlan object. \
Be thorough — populate EVERY available field from the source content.

═══════════════════════════════════════════════════════
FIELD-BY-FIELD EXTRACTION RULES
═══════════════════════════════════════════════════════

── name ──
Short, clean task name. Strip block prefixes if redundant (e.g. "Priority block: " can stay).
For a task listed as "6:30 AM — Breathing / Meditation", name = "Breathing / Meditation".

── timing ──
The clock time or time-of-day label. Examples: "5:40 AM", "Morning", "Evening", "Post-workout".
Look for time columns, "When" columns, or time prefixes in the task name (e.g. "6:30 AM — ...").

── target_value ──
CRITICAL for brief_today blocks: copy the FULL text from the "Week 1 easy start" column VERBATIM.
This contains the exercise list for the block (e.g. "2 rounds: band row, wall slide, dead bug/bird dog").
For other tasks: dosage, reps, duration, or measurable goal.

── description ──
The WHY — benefits and reasons. Format as one benefit per line:
  "Reduces stress
  Improves focus"
Use "Why" or "Benefits" column content.

── how_to ──
CRITICAL — Always populate when steps exist.

  For brief_today WORKOUT BLOCKS: give a short block-level instruction only.
  Do NOT list individual exercises here — those come from "Week 1 easy start" / target_value.
  Example: "Use controlled tempo. Pull elbows back toward ribs. Keep core braced throughout."

  For BREATHING / MEDITATION: one step per line.
  For SUPPLEMENTS: exact instructions (take with, avoid with, timing).
  For HABITS: action steps.

── safety_notes ──
Pain rules, contraindications, stop conditions.
For brief_today: use "If behind schedule" column content here.

── why_mechanism ──
Deep scientific mechanism. Only if source has a "mechanism" or "science" column. Otherwise null.

── schedule ──
Normalize to: 'daily', '2x/week', '3x/week', 'Mon/Wed/Fri', 'Mon/Thu', 'weekly', 'as needed'.
Default to 'daily' for brief_today and supplement items with no explicit schedule.

── pillar ──
Normalize sheet name: lowercase, spaces→underscores, strip leading number+underscore prefix.
Examples: "02_Brief_Today" → "brief_today", "09_Supplements" → "supplements".

── is_reference ──
Set True ONLY for: nutrition, sleep_recovery, cognitive_social, exercise_library.
Everything else is False — brief_today, supplements, etc. are actionable daily tasks.

── video_link / link ──
Only populate with actual URLs (starting with http). Leave null for search terms / empty cells.

═══════════════════════════════════════════════════════
CONTENT SOURCES  (extract tasks from these sheets only)
═══════════════════════════════════════════════════════
- brief_today sheet: daily schedule blocks — copy "Week 1 easy start" into target_value verbatim
- exercise_library sheet: extract as individual reference tasks (is_reference=True)
- supplements sheet: dosage → target_value, instructions → how_to, benefits → description
- nutrition, sleep_recovery, cognitive_social sheets: is_reference=True

Do not invent data. Only extract what is present in the source.
Do not emit tasks for sheets not listed above.
"""


def normalize_with_llm(content: str) -> IngestedPlan:
    """Legacy / JSON path: send all content to LLM and get back IngestedPlan."""
    from langchain_core.messages import HumanMessage, SystemMessage

    llm = _make_llm(max_tokens=32000)
    structured_llm = llm.with_structured_output(IngestedPlan)

    result = structured_llm.invoke([
        SystemMessage(content=_SYSTEM_PROMPT),
        HumanMessage(content=f"Extract the plan from this content:\n\n{content}"),
    ])
    return result  # type: ignore[return-value]


# keep old name as alias for any callers that imported it directly
normalize_with_claude = normalize_with_llm


# ── DB helpers ────────────────────────────────────────────────────────────

def _wipe_user_data(db: Session, user_id: str) -> date | None:
    prev_plan = (
        db.query(Plan)
        .filter(Plan.user_id == user_id, Plan.is_active == True)  # noqa: E712
        .order_by(Plan.uploaded_at.desc())
        .first()
    )
    prev_start = prev_plan.rotation_start_date if prev_plan else None

    db.query(DailyTodo).filter(DailyTodo.user_id == user_id).delete()
    for plan in db.query(Plan).filter(Plan.user_id == user_id).all():
        db.delete(plan)
    db.flush()
    return prev_start


def _save_tasks(db: Session, plan_id: str, user_id: str, tasks: list[IngestedTask]) -> int:
    for t in tasks:
        db.add(TaskTemplate(
            id=str(uuid.uuid4()),
            plan_id=plan_id,
            user_id=user_id,
            pillar=t.pillar,
            name=t.name,
            description=t.description,
            schedule=t.schedule,
            timing=t.timing,
            target_value=t.target_value,
            unit=t.unit,
            how_to=t.how_to,
            why_mechanism=t.why_mechanism,
            safety_notes=t.safety_notes,
            video_link=t.video_link,
            link=t.link,
            benefit_tags=t.benefit_tags,
            source_key=t.source_key,
            is_reference=t.is_reference,
            extra_metadata=json.dumps(t.extra_metadata or {}),
        ))
    return len(tasks)


def _save_rotation(db: Session, plan_id: str, user_id: str, days: list[IngestedRotationDay]) -> int:
    for rd in days:
        db.add(RotationDay(
            id=str(uuid.uuid4()),
            plan_id=plan_id,
            user_id=user_id,
            day_number=rd.day_number,
            week_number=rd.week_number,
            block_name=rd.block_name,
            morning_time=rd.morning_time,
            warm_up_min=rd.warm_up_min,
            upper_back_core_min=rd.upper_back_core_min,
            secondary_min=rd.secondary_min,
            cool_down_min=rd.cool_down_min,
            total_min=rd.total_min,
            fits_60=rd.fits_60,
            priority_exercises=rd.priority_exercises,
            secondary_exercises=rd.secondary_exercises,
            week_rule=rd.week_rule,
            notes=rd.notes,
            warm_up=rd.warm_up,
            priority_block=rd.priority_block,
            secondary_block=rd.secondary_block,
            cardio_steps=rd.cardio_steps,
            cool_down=rd.cool_down,
            nutrition_focus=rd.nutrition_focus,
            intensity_cap=rd.intensity_cap,
        ))
    return len(days)


def _save_screenings(db: Session, plan_id: str, user_id: str, items: list[IngestedScreening]) -> int:
    for s in items:
        db.add(Screening(
            id=str(uuid.uuid4()),
            plan_id=plan_id,
            user_id=user_id,
            pillar=s.pillar,
            name=s.name,
            description=s.description,
            frequency_months=s.frequency_months,
            target_value=s.target_value,
        ))
    return len(items)


def prefill_todos(db: Session, plan_id: str, user_id: str) -> int:
    """Create DailyTodo rows for next 30 days for all non-reference templates."""
    templates = (
        db.query(TaskTemplate)
        .filter(TaskTemplate.plan_id == plan_id, TaskTemplate.is_reference == False)  # noqa: E712
        .all()
    )

    today = date.today()
    count = 0
    for offset in range(30):
        target = today + timedelta(days=offset)
        for tmpl in templates:
            if _is_scheduled_today(tmpl.schedule, target):
                db.add(DailyTodo(
                    id=str(uuid.uuid4()),
                    template_id=tmpl.id,
                    user_id=user_id,
                    date=target,
                    completed=False,
                ))
                count += 1
    return count


# ── Questionnaire-driven ingest (no LLM, no xlsx upload) ─────────────────

def ingest_from_workbook_json(
    workbook_json: dict,
    db: Session,
    user_id: str = DEFAULT_USER_ID,
) -> dict:
    """
    Create a Plan + TaskTemplates + RotationDays + Screenings + DailyTodos from
    questionnaire-generated workbook_json. No LLM is needed — all data is deterministic.

    Deactivates any existing active plans for *user_id* before creating the new one.
    Returns {"plan_id": str, "plan_name": str}.
    """
    profile = workbook_json.get("user_profile", {})
    plan_name = f"{profile.get('name', 'User')}'s Longevity Plan"

    # Recover rotation_start_date from the current active plan before wiping it.
    # This preserves the existing rotation cycle position so the user's weekly
    # rotation day count is not reset just because they regenerated their plan.
    prev_plan = (
        db.query(Plan)
        .filter(Plan.user_id == user_id, Plan.is_active == True)  # noqa: E712
        .order_by(Plan.uploaded_at.desc())
        .first()
    )
    prev_rotation_start = prev_plan.rotation_start_date if prev_plan else None

    # Delete stale DailyTodo rows before creating the new plan.
    # generate_todos_for_date checks for existing DailyTodos FIRST and returns them
    # immediately — so simply deactivating the old plan is not enough; stale rows
    # pointing at old TaskTemplate IDs would keep leaking through.
    db.query(DailyTodo).filter(DailyTodo.user_id == user_id).delete()

    # Deactivate existing plans so only the new plan is marked is_active=True.
    db.query(Plan).filter(
        Plan.user_id == user_id, Plan.is_active == True  # noqa: E712
    ).update({"is_active": False})
    db.flush()

    plan = Plan(
        name=plan_name,
        user_id=user_id,
        plan_json=json.dumps(workbook_json),
        is_active=False,
        status="draft",
        rotation_start_date=prev_rotation_start,
    )
    db.add(plan)
    db.flush()  # populate plan.id before relationships

    # TaskTemplates
    for task_data in workbook_json.get("tasks", []):
        db.add(TaskTemplate(
            plan_id=plan.id,
            user_id=user_id,
            pillar=task_data.get("pillar", "general"),
            name=task_data.get("name", ""),
            description=task_data.get("description"),
            schedule=task_data.get("schedule"),
            timing=task_data.get("timing"),
            target_value=task_data.get("target_value"),
            benefit_tags=task_data.get("benefit_tags"),
            source_key=task_data.get("source_key"),
            safety_notes=task_data.get("safety_notes"),
            how_to=task_data.get("how_to"),
            why_mechanism=task_data.get("why_mechanism"),
            is_reference=task_data.get("is_reference", False),
        ))

    # RotationDays
    for rd_data in workbook_json.get("rotation_days", []):
        db.add(RotationDay(
            plan_id=plan.id,
            user_id=user_id,
            day_number=rd_data.get("day_number", 1),
            week_number=rd_data.get("week_number"),
            block_name=rd_data.get("block_name", ""),
            morning_time=rd_data.get("morning_time"),
            warm_up_min=rd_data.get("warm_up_min"),
            upper_back_core_min=rd_data.get("upper_back_core_min"),
            secondary_min=rd_data.get("secondary_min"),
            cool_down_min=rd_data.get("cool_down_min"),
            total_min=rd_data.get("total_min"),
            fits_60=rd_data.get("fits_60"),
            priority_exercises=rd_data.get("priority_exercises"),
            secondary_exercises=rd_data.get("secondary_exercises"),
            week_rule=rd_data.get("week_rule"),
        ))

    # Screenings
    for sc_data in workbook_json.get("screenings", []):
        db.add(Screening(
            plan_id=plan.id,
            user_id=user_id,
            pillar=sc_data.get("pillar", "screenings_safety"),
            name=sc_data.get("name", ""),
            description=sc_data.get("description"),
            frequency_months=sc_data.get("frequency_months"),
            target_value=sc_data.get("target_value"),
        ))

    db.flush()

    # Pre-generate DailyTodo rows for next 30 days (matches run_ingest behaviour)
    prefill_todos(db, plan.id, user_id)

    db.commit()
    db.refresh(plan)
    return {"plan_id": plan.id, "plan_name": plan_name}


def _loads_or(raw: str | None, fallback: object) -> object:
    """Parse a stored JSON string; return *fallback* on None/empty/invalid JSON."""
    if not raw:
        return fallback
    try:
        return json.loads(raw)
    except (ValueError, TypeError):
        return fallback


def _plan_to_json(db: Session, plan_id: str) -> dict:
    """
    Serialize the already-persisted DB rows for *plan_id* back into the same
    dict shape ``ingest_from_workbook_json`` consumes, so uploaded plans get a
    lossless canonical ``plan_json`` (single source of truth for edit/review).

    Pure and deterministic: rows are ordered by primary key so repeated calls
    produce identical output. ``extra_metadata`` / ``exercises_json`` are parsed
    back into structures (dict / list) rather than left as raw strings.

    MODIFY_WORKSHEET_PLAN_FINAL §4.1 / §6. The DB ``status`` column is NOT read
    or written here — only the inert ``plan_status`` JSON field is emitted.
    """
    plan = db.query(Plan).filter(Plan.id == plan_id).one()

    tasks = (
        db.query(TaskTemplate)
        .filter(TaskTemplate.plan_id == plan_id)
        .order_by(TaskTemplate.id)
        .all()
    )
    rotation_days = (
        db.query(RotationDay)
        .filter(RotationDay.plan_id == plan_id)
        .order_by(RotationDay.day_number, RotationDay.id)
        .all()
    )
    screenings = (
        db.query(Screening)
        .filter(Screening.plan_id == plan_id)
        .order_by(Screening.id)
        .all()
    )

    tasks_json = [
        {
            "task_id": t.id,
            "pillar": t.pillar,
            "name": t.name,
            "description": t.description,
            "schedule": t.schedule,
            "timing": t.timing,
            "target_value": t.target_value,
            "unit": t.unit,
            "benefit_tags": t.benefit_tags,
            "source_key": t.source_key,
            "link": t.link,
            "video_link": t.video_link,
            "safety_notes": t.safety_notes,
            "how_to": t.how_to,
            "why_mechanism": t.why_mechanism,
            "is_reference": t.is_reference,
            "extra_metadata": _loads_or(t.extra_metadata, {}),
            "exercises_json": _loads_or(t.exercises_json, []),
        }
        for t in tasks
    ]

    rotation_json = [
        {
            "day_number": rd.day_number,
            "week_number": rd.week_number,
            "block_name": rd.block_name,
            "warm_up": rd.warm_up,
            "priority_block": rd.priority_block,
            "secondary_block": rd.secondary_block,
            "cardio_steps": rd.cardio_steps,
            "cool_down": rd.cool_down,
            "nutrition_focus": rd.nutrition_focus,
            "intensity_cap": rd.intensity_cap,
            "source_key": rd.source_key,
            "sets": rd.sets,
            "reps": rd.reps,
            "duration": rd.duration,
            "notes": rd.notes,
            "morning_time": rd.morning_time,
            "warm_up_min": rd.warm_up_min,
            "upper_back_core_min": rd.upper_back_core_min,
            "secondary_min": rd.secondary_min,
            "cool_down_min": rd.cool_down_min,
            "total_min": rd.total_min,
            "fits_60": rd.fits_60,
            "priority_exercises": rd.priority_exercises,
            "secondary_exercises": rd.secondary_exercises,
            "week_rule": rd.week_rule,
            "extra_metadata": _loads_or(rd.extra_metadata, {}),
        }
        for rd in rotation_days
    ]

    screenings_json = [
        {
            "pillar": sc.pillar,
            "name": sc.name,
            "description": sc.description,
            "frequency_months": sc.frequency_months,
            "target_value": sc.target_value,
        }
        for sc in screenings
    ]

    return {
        "format_version": "upload_v2",
        "plan_status": "draft",
        "review": {"auto_removed": [], "flags": [], "advisor_notes": []},
        "user_profile": {"name": plan.name},
        "tasks": tasks_json,
        "rotation_days": rotation_json,
        "screenings": screenings_json,
    }


# ── Public entry point ────────────────────────────────────────────────────

class IngestResult(BaseModel):
    plan_id: str
    plan_name: str
    tasks_imported: int
    rotation_days_imported: int
    screenings_imported: int
    todos_prefilled: int
    pillars_found: list[str]


def run_ingest(
    db: Session,
    file_content: bytes,
    filename: str,
    rotation_start_date: date | None = None,
) -> IngestResult:
    """Full pipeline: extract → per-sheet LLM → save DB → enrich exercises → prefill todos."""
    user_id = DEFAULT_USER_ID
    is_json = filename.lower().endswith('.json')

    # 1. Python-direct: exercise library (no LLM)
    exercise_library: dict[str, dict] = {}
    if not is_json:
        exercise_library = _read_exercise_library(file_content)
        print(f"[ingest] exercise_library: {len(exercise_library)} exercises")

    # 2. Python-direct: screenings and rotation (always deterministic)
    all_screenings: list[IngestedScreening] = []
    all_rotation: list[IngestedRotationDay] = []
    if not is_json:
        all_screenings = _read_screenings_from_excel(file_content)
        print(f"[ingest] screenings: {len(all_screenings)}")
        all_rotation = _read_rotation_from_excel(file_content)
        print(f"[ingest] rotation days: {len(all_rotation)}")

    # 3. LLM: extract tasks one sheet at a time so no sheet is ever dropped
    plan_name = "Personal Longevity & Fitness Plan"
    valid_tasks: list[IngestedTask] = []

    if is_json:
        content_text = extract_json_text(io.BytesIO(file_content))
        ingested = normalize_with_llm(content_text)
        valid_tasks = [t for t in ingested.tasks if t.pillar not in SKIP_SHEETS]
        plan_name = ingested.plan_name or filename
    else:
        llm = _make_llm()
        sheet_texts = _extract_sheet_texts(file_content)
        for pillar, sheet_text in sheet_texts.items():
            if pillar == 'brief_today':
                tasks = _read_brief_today_from_excel(file_content)
                print(f"[ingest] {pillar}: {len(tasks)} tasks (direct)")
                valid_tasks.extend(tasks)
                continue
            if pillar == 'supplements':
                tasks = _read_supplements_from_excel(file_content)
                print(f"[ingest] {pillar}: {len(tasks)} tasks (direct)")
                valid_tasks.extend(tasks)
                continue
            tasks = _extract_sheet_tasks(llm, sheet_text, pillar)
            print(f"[ingest] {pillar}: {len(tasks)} tasks")
            valid_tasks.extend(tasks)

    # 4. Wipe old data; recover previous rotation start date
    prev_start = _wipe_user_data(db, user_id)
    effective_start = rotation_start_date or prev_start

    # 5. Create new Plan row (draft until user approves via /activate)
    plan_id = str(uuid.uuid4())
    db.add(Plan(
        id=plan_id,
        name=plan_name,
        is_active=False,
        status="draft",
        user_id=user_id,
        rotation_start_date=effective_start,
        plan_json='{}',  # backfilled with the canonical JSON in step 9 below
    ))
    db.flush()

    # 6. Save tasks, rotation days, screenings
    tasks_n = _save_tasks(db, plan_id, user_id, valid_tasks)
    rotation_n = _save_rotation(db, plan_id, user_id, all_rotation)
    screening_n = _save_screenings(db, plan_id, user_id, all_screenings)
    db.flush()

    # 7. Enrich brief_today blocks with per-exercise details from the library
    _enrich_block_exercises(db, plan_id, exercise_library)

    # 8. Pre-generate 30 days of DailyTodo rows
    todos_n = prefill_todos(db, plan_id, user_id)

    # 9. Backfill canonical plan_json from the now-saved rows so uploaded plans
    #    are editable through the same JSON-as-source-of-truth path as
    #    questionnaire plans (MODIFY_WORKSHEET_PLAN_FINAL §4.1).
    plan = db.query(Plan).filter(Plan.id == plan_id).one()
    plan.plan_json = json.dumps(_plan_to_json(db, plan_id))

    db.commit()

    return IngestResult(
        plan_id=plan_id,
        plan_name=plan_name,
        tasks_imported=tasks_n,
        rotation_days_imported=rotation_n,
        screenings_imported=screening_n,
        todos_prefilled=todos_n,
        pillars_found=sorted({t.pillar for t in valid_tasks}),
    )
