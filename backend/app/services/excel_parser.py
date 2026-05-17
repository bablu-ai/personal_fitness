"""
Excel parser — open architecture.
Sheet names become pillar names dynamically (numeric prefix stripped).
Header row is detected automatically (widest row in first 15 rows).
Column headers matched flexibly; unknown columns stored as JSON metadata.
"""
import json
import re
import uuid
from typing import BinaryIO
import openpyxl
import re as _re
from app.constants import COLUMN_HINTS, SKIP_SHEETS, SCREENING_SHEETS, REFERENCE_PILLARS


def _normalize_pillar(sheet_name: str) -> str:
    """Strip leading numeric prefix (e.g. '09_Supplements' → 'supplements')."""
    name = sheet_name.strip()
    name = re.sub(r'^\d+_', '', name)   # remove "09_" prefix
    return name.lower().replace(" ", "_")


def _find_header_row(rows: list[tuple]) -> int:
    """
    Return the index of the header row.
    Strategy: the header row is the row with the most non-None cells
    within the first 15 rows. Ties broken by earliest row.
    """
    best_idx, best_count = 0, 0
    for i, row in enumerate(rows[:15]):
        count = sum(1 for c in row if c is not None)
        if count > best_count:
            best_count, best_idx = count, i
    return best_idx


def _is_section_header(row: tuple, name_col_idx: int | None) -> bool:
    """
    Detect rows that are section headers / dividers, not real data.
    A row is a section header if:
    - Only 1–2 cells are populated, OR
    - The name column value is ALL CAPS (e.g. "CORE TARGETS")
    """
    populated = [c for c in row if c is not None]
    if len(populated) <= 1:
        return True
    if name_col_idx is not None:
        val = row[name_col_idx]
        if val and isinstance(val, str) and val.strip() == val.strip().upper() and len(val.strip()) > 2:
            return True
    return False


def _map_column(header: str) -> str | None:
    """Return canonical DB field name for a column header, or None if unknown."""
    normalized = header.strip().lower()
    for field, hints in COLUMN_HINTS.items():
        if normalized in hints:
            return field
    return None


def parse_workbook(file: BinaryIO, plan_id: str, user_id: str = "default") -> list[dict]:
    """
    Parse an .xlsx workbook.
    - Each worksheet → one pillar (numeric prefix stripped from sheet name)
    - Header row auto-detected (widest row in first 15 rows)
    - Each data row → one task_template dict ready for DB insert
    - Section header rows within the sheet are skipped automatically
    """
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    tasks: list[dict] = []

    for sheet_name in wb.sheetnames:
        pillar = _normalize_pillar(sheet_name)

        if pillar in SKIP_SHEETS:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            continue

        header_idx = _find_header_row(rows)
        raw_headers = rows[header_idx]

        headers = [
            str(h).strip() if h is not None else f"col_{i}"
            for i, h in enumerate(raw_headers)
        ]
        column_map = {i: _map_column(h) for i, h in enumerate(headers)}

        # Find which column index maps to "name" so we can detect section headers
        name_col_idx = next(
            (i for i, field in column_map.items() if field == "name"), None
        )

        for row in rows[header_idx + 1:]:
            if not any(cell is not None for cell in row):
                continue

            if _is_section_header(row, name_col_idx):
                continue

            task: dict = {
                "id": str(uuid.uuid4()),
                "plan_id": plan_id,
                "user_id": user_id,
                "pillar": pillar,
                "extra_metadata": {},
            }

            for i, value in enumerate(row):
                if value is None:
                    continue
                field = column_map.get(i)
                str_value = str(value).strip()
                if not str_value:
                    continue
                if field:
                    # Don't overwrite if already set (first mapped column wins)
                    if field not in task:
                        task[field] = str_value
                else:
                    task["extra_metadata"][headers[i]] = str_value

            if not task.get("name"):
                continue

            task["is_reference"] = pillar in REFERENCE_PILLARS
            task["extra_metadata"] = json.dumps(task["extra_metadata"])
            tasks.append(task)

    return tasks


def _parse_frequency_months(schedule: str | None) -> int | None:
    """Convert a frequency string to months. Returns None if unparseable."""
    if not schedule:
        return None
    s = schedule.strip().lower()
    # "annual" / "1 yr" / "yearly"
    if any(k in s for k in ("annual", "yearly", "1 yr", "1yr", "per year", "every year")):
        return 12
    # "2 yrs" / "biennial"
    m = _re.search(r'(\d+)\s*yr', s)
    if m:
        return int(m.group(1)) * 12
    # "every N months" / "N months"
    m = _re.search(r'(\d+)\s*month', s)
    if m:
        return int(m.group(1))
    # "quarterly"
    if "quarter" in s:
        return 3
    # "biannual" / "twice a year" / "6 months"
    if any(k in s for k in ("biannual", "twice a year", "semi-annual", "6 month")):
        return 6
    # "monthly"
    if "monthly" in s or "every month" in s:
        return 1
    # "every 5 years" style (already caught by yr regex above, but just in case)
    m = _re.search(r'every\s+(\d+)', s)
    if m:
        return int(m.group(1)) * 12
    return None


def parse_screening_sheets(file: BinaryIO, plan_id: str, user_id: str = "default") -> list[dict]:
    """
    Parse blood_markers + screenings_safety sheets into Screening dicts.
    Frequency is extracted from the 'schedule' / 'frequency' column.
    """
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    screenings: list[dict] = []

    for sheet_name in wb.sheetnames:
        pillar = _normalize_pillar(sheet_name)
        if pillar not in SCREENING_SHEETS:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        header_idx = _find_header_row(rows)
        raw_headers = rows[header_idx]
        headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(raw_headers)]
        column_map = {i: _map_column(h) for i, h in enumerate(headers)}
        name_col_idx = next((i for i, f in column_map.items() if f == "name"), None)

        for row in rows[header_idx + 1:]:
            if not any(cell is not None for cell in row):
                continue
            if _is_section_header(row, name_col_idx):
                continue

            entry: dict = {
                "id": str(uuid.uuid4()),
                "plan_id": plan_id,
                "user_id": user_id,
                "pillar": pillar,
                "extra_metadata": {},
            }

            for i, value in enumerate(row):
                if value is None:
                    continue
                field = column_map.get(i)
                str_value = str(value).strip()
                if not str_value:
                    continue
                if field:
                    if field not in entry:
                        entry[field] = str_value
                else:
                    entry["extra_metadata"][headers[i]] = str_value

            if not entry.get("name"):
                continue

            entry["frequency_months"] = _parse_frequency_months(entry.get("schedule"))
            entry["extra_metadata"] = json.dumps(entry["extra_metadata"])
            screenings.append(entry)

    return screenings


def get_pillars_from_workbook(file: BinaryIO) -> list[str]:
    """Return pillar names from sheet names (numeric prefix stripped, skip-sheets excluded)."""
    wb = openpyxl.load_workbook(file, read_only=True)
    return [
        _normalize_pillar(s) for s in wb.sheetnames
        if _normalize_pillar(s) not in SKIP_SHEETS
    ]


# Column hints specific to the rotation sheet.
# v3 columns: Day, Week, Type, Warm-up, Priority block, Secondary block,
#             Cardio / steps, Sets, Reps, Duration, Cool-down, Nutrition focus,
#             Intensity cap, Milestone / notes, Source key
# v4 columns: Day, Week, Focus, Morning time, Warm-up min, Upper back + core min,
#             Secondary min, Cool-down min, Total min, Fits 60?,
#             Priority exercises, Secondary exercises, Week rule, Notes
_ROTATION_HINTS: dict[str, list[str]] = {
    "day_number":           ["day", "day #", "day#", "day number", "#"],
    "week_number":          ["week", "wk"],
    # v3: 'Type'; v4: 'Focus' — both map to block_name (required field)
    "block_name":           ["focus", "type", "block", "exercise", "workout", "activity",
                             "name", "movement", "session", "session type"],
    # v4 time-budget fields
    "morning_time":         ["morning time"],
    "warm_up_min":          ["warm-up min", "warmup min", "warm up min"],
    "upper_back_core_min":  ["upper back + core min", "upper back+core min",
                             "upper back & core min", "upper back/core min"],
    "secondary_min":        ["secondary min"],
    "cool_down_min":        ["cool-down min", "cooldown min", "cool down min"],
    "total_min":            ["total min", "total"],
    "fits_60":              ["fits 60?", "fits 60", "fits?"],
    "priority_exercises":   ["priority exercises"],
    "secondary_exercises":  ["secondary exercises"],
    "week_rule":            ["week rule", "week 1 rule", "rule"],
    # v3 legacy hints (no-op for v4 but retained for backward compatibility)
    "warm_up":              ["warm-up", "warm up", "warmup"],
    "priority_block":       ["priority block", "priority block: upper back + core",
                             "priority block: upper back", "priority block: lower body",
                             "priority block: full body", "priority"],
    "secondary_block":      ["secondary block", "secondary", "accessory"],
    "cardio_steps":         ["cardio / steps", "cardio/steps", "cardio", "steps"],
    "cool_down":            ["cool-down", "cool down", "cooldown"],
    "nutrition_focus":      ["nutrition focus", "nutrition", "peri-workout nutrition"],
    "intensity_cap":        ["intensity cap", "intensity", "rpe", "effort"],
    "source_key":           ["source key", "source", "source_key"],
    "sets":                 ["sets", "set"],
    "reps":                 ["reps", "rep", "repetitions"],
    "duration":             ["duration", "minutes", "mins"],
    "notes":                ["milestone / notes", "milestone/notes", "notes", "note",
                             "milestones", "milestone"],
}


def _map_rotation_column(header: str) -> str | None:
    normalized = header.strip().lower()
    for field, hints in _ROTATION_HINTS.items():
        if normalized in hints:
            return field
    return None


def parse_rotation_sheet(file: BinaryIO, plan_id: str, user_id: str = "default") -> list[dict]:
    """
    Parse the 30-day rotation sheet into RotationDay dicts.
    Looks for a sheet whose normalized name is '30day_rotation'.
    Returns [] if the sheet is absent.
    """
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)

    rotation_sheet = None
    for sheet_name in wb.sheetnames:
        if _normalize_pillar(sheet_name) == "30day_rotation":
            rotation_sheet = wb[sheet_name]
            break

    if rotation_sheet is None:
        return []

    rows = list(rotation_sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    header_idx = _find_header_row(rows)
    raw_headers = rows[header_idx]
    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(raw_headers)]
    column_map = {i: _map_rotation_column(h) for i, h in enumerate(headers)}

    name_col_idx = next((i for i, f in column_map.items() if f == "block_name"), None)

    rotation_days: list[dict] = []
    day_counter = 0

    for row in rows[header_idx + 1:]:
        if not any(cell is not None for cell in row):
            continue
        if _is_section_header(row, name_col_idx):
            continue

        entry: dict = {
            "id": str(uuid.uuid4()),
            "plan_id": plan_id,
            "user_id": user_id,
            "extra_metadata": {},
        }

        for i, value in enumerate(row):
            if value is None:
                continue
            field = column_map.get(i)
            str_value = str(value).strip()
            if not str_value:
                continue
            if field:
                if field not in entry:
                    entry[field] = str_value
            else:
                entry["extra_metadata"][headers[i]] = str_value

        if not entry.get("block_name"):
            continue

        # If no day_number column, assign sequentially
        if "day_number" not in entry:
            day_counter += 1
            entry["day_number"] = day_counter
        else:
            try:
                entry["day_number"] = int(float(entry["day_number"]))
                day_counter = entry["day_number"]
            except (ValueError, TypeError):
                day_counter += 1
                entry["day_number"] = day_counter

        entry["extra_metadata"] = json.dumps(entry["extra_metadata"])
        rotation_days.append(entry)

    return rotation_days
