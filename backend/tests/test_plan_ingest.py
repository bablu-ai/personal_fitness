import io
import json

import openpyxl

from app.db.models import Plan, RotationDay, Screening, TaskTemplate
from app.services import plan_ingest
from app.services.plan_ingest import (
    IngestedPlan,
    IngestedTask,
    _plan_to_json,
    _read_brief_today_from_excel,
    _read_exercise_library,
    _read_supplements_from_excel,
    run_ingest,
)


def _xlsx_bytes(rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "02_Brief_Today"
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_brief_today_v5_split_exercise_columns_and_minimums():
    content = _xlsx_bytes([
        ["Brief Today"],
        [None],
        ["Start time", "End time", "Exercise 1", "Exercise 2", "Exercise 3", "Dose / target", "Progression", "Why it matters", "Notes"],
        ["05:40", "05:50", "Band row", "Wall slide", "Dead bug", "2 rounds", "+1 set", "Posture/core", "Keep this block"],
        [None],
        ["THIS WEEK -- NON-NEGOTIABLE MINIMUMS"],
        ["Pillar", "Minimum", "Week 1", "Week 2", "Track", "Notes"],
        ["Upper back", "3 sessions/week", "Band row", "+1 set", "Sessions", "Start here"],
    ])

    tasks = _read_brief_today_from_excel(content)

    assert len(tasks) == 2
    assert tasks[0].name == "Band row + Wall slide + Dead bug"
    assert tasks[0].timing == "05:40-05:50"
    assert tasks[0].target_value == "2 rounds"
    assert tasks[0].extra_metadata == {
        "exercise_names": "Band row; Wall slide; Dead bug",
        "progression": "+1 set",
        "notes": "Keep this block",
    }
    assert tasks[1].name == "(must) Upper back"
    assert tasks[1].schedule == "weekly"
    assert tasks[1].extra_metadata["must"] == "true"


def _exercise_library_xlsx_bytes(youtube_formula: str, gif_formula: str) -> bytes:
    """Build a minimal 07_Exercise_Library xlsx with one exercise row."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "07_Exercise_Library"
    ws.append([
        "Exercise", "Category", "Setup", "Starting Position",
        "Step-by-step", "Core/Bracing", "Common mistakes",
        "Week 1 dosage", "Safety stop", "Why it matters",
        "YouTube Demo", "Animated GIF",
    ])
    ws.append([
        "Cat-cow", "Warm-up", None, None,
        None, None, None,
        None, None, None,
        youtube_formula, gif_formula,
    ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_exercise_library_hyperlink_formula_extracts_real_url():
    """=HYPERLINK("url","display text") cells must yield the real URL, not the display text."""
    yt_url = "https://www.youtube.com/results?search_query=Cat-cow+exercise+proper+form+demonstration+older+adult+safe"
    gif_url = "https://giphy.com/search/cat-cow-stretch"
    content = _exercise_library_xlsx_bytes(
        f'=HYPERLINK("{yt_url}","YouTube demo search")',
        f'=HYPERLINK("{gif_url}","GIF demo")',
    )
    library = _read_exercise_library(content)

    assert "cat-cow" in library
    ex = library["cat-cow"]
    assert ex["video_link"] == yt_url, (
        f"Expected real YouTube URL but got: {ex['video_link']!r}\n"
        "This means the HYPERLINK formula URL was not extracted — likely data_only=True is back."
    )
    assert ex["gif_link"] == gif_url


def test_exercise_library_display_text_not_used_as_search_term():
    """Regression: display text like 'YouTube demo search' must NOT become the search query."""
    yt_url = "https://www.youtube.com/results?search_query=Cat-cow+exercise+proper+form"
    content = _exercise_library_xlsx_bytes(
        f'=HYPERLINK("{yt_url}","YouTube demo search")',
        None,
    )
    library = _read_exercise_library(content)

    ex = library.get("cat-cow", {})
    video_link = ex.get("video_link", "")
    assert "YouTube+demo+search" not in (video_link or ""), (
        f"Display text leaked into search query: {video_link!r}"
    )


def test_supplements_parser_extracts_metadata_and_schedule():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "09_Supplements"
    ws.append(["Supplements"])
    ws.append([None])
    ws.append([
        "Supplement", "Cat.", "Status", "Dose", "Frequency", "Timing / food",
        "Trigger marker", "Skip if", "Stop rule", "Do NOT combine with",
        "Why (mechanism)", "Evidence", "Source key",
    ])
    ws.append([
        "Creatine monohydrate", "B", "Active", "5 g", "DAILY (no exceptions)",
        "Any time", "N/A", "Never skip", "Stop if kidney disease develops",
        "Nothing major", "ATP regeneration", "STRONG", "KREIDER-2017",
    ])
    ws.append([
        "Vitamin B12", "A", "Discuss", "500 mcg", "EOD or 2×/wk if normal",
        "AM", "Serum B12 + MMA", "B12 >500", "Stop full dose when MMA normalizes",
        "Nothing", "Absorption declines with age", "STRONG", "USPSTF-B12",
    ])
    buf = io.BytesIO()
    wb.save(buf)

    tasks = _read_supplements_from_excel(buf.getvalue())
    assert len(tasks) == 2
    assert tasks[0].name == "Creatine monohydrate"
    assert tasks[0].schedule == "daily"
    assert tasks[0].extra_metadata == {
        "status": "Active",
        "category": "B",
        "trigger_marker": "N/A",
        "skip_if": "Never skip",
        "stop_rule": "Stop if kidney disease develops",
        "do_not_combine_with": "Nothing major",
        "evidence": "STRONG",
    }
    assert tasks[1].schedule == "every other day"


# ── _plan_to_json / run_ingest canonical-JSON backfill (§4.1) ───────────────


def _seed_plan(db) -> str:
    """Insert a Plan with two tasks, a rotation day and a screening; return plan_id."""
    plan = Plan(id="plan-1", name="Seed Plan", user_id="default", plan_json="{}")
    db.add(plan)
    db.flush()
    db.add(TaskTemplate(
        id="task-a", plan_id="plan-1", user_id="default",
        pillar="brief_today", name="Band row", schedule="daily",
        extra_metadata=json.dumps({"must": "true"}),
        exercises_json=json.dumps([{"name": "Band row", "sets": "3"}]),
    ))
    db.add(TaskTemplate(
        id="task-b", plan_id="plan-1", user_id="default",
        pillar="supplements", name="Creatine", schedule="daily",
        extra_metadata=None, exercises_json=None,  # NULL → must become {} / []
    ))
    db.add(RotationDay(
        id="rd-1", plan_id="plan-1", user_id="default",
        day_number=1, week_number=1, block_name="Upper body",
        priority_exercises="Band row, Wall slide",
    ))
    db.add(Screening(
        id="sc-1", plan_id="plan-1", user_id="default",
        pillar="blood_markers", name="ApoB", frequency_months=12,
    ))
    db.flush()
    return "plan-1"


def test_plan_to_json_round_trips_db_rows(db_session):
    plan_id = _seed_plan(db_session)

    result = _plan_to_json(db_session, plan_id)

    assert result["format_version"] == "upload_v2"
    assert result["plan_status"] == "draft"
    assert result["review"] == {"auto_removed": [], "flags": [], "advisor_notes": []}
    assert result["user_profile"] == {"name": "Seed Plan"}

    assert len(result["tasks"]) == 2
    assert len(result["rotation_days"]) == 1
    assert len(result["screenings"]) == 1

    for task in result["tasks"]:
        assert task["task_id"]  # non-empty stable identity
    by_id = {t["task_id"]: t for t in result["tasks"]}
    assert set(by_id) == {"task-a", "task-b"}

    # extra_metadata / exercises_json parsed back to structures, not raw strings
    assert by_id["task-a"]["extra_metadata"] == {"must": "true"}
    assert by_id["task-a"]["exercises_json"] == [{"name": "Band row", "sets": "3"}]
    # NULL columns fall back to {} / []
    assert by_id["task-b"]["extra_metadata"] == {}
    assert by_id["task-b"]["exercises_json"] == []

    assert result["rotation_days"][0]["day_number"] == 1
    assert result["rotation_days"][0]["priority_exercises"] == "Band row, Wall slide"
    assert result["screenings"][0]["name"] == "ApoB"

    # whole structure must be JSON-serializable
    json.dumps(result)


def test_run_ingest_stores_non_empty_plan_json(db_session, monkeypatch):
    """run_ingest must backfill a real plan_json that parses to the DB task count."""
    fake_plan = IngestedPlan(
        plan_name="Uploaded Plan",
        tasks=[
            IngestedTask(pillar="brief_today", name="Band row", schedule="daily"),
            IngestedTask(pillar="supplements", name="Creatine", schedule="daily"),
        ],
    )
    monkeypatch.setattr(plan_ingest, "normalize_with_llm", lambda _content: fake_plan)

    result = run_ingest(db_session, b'{"tasks": []}', "plan.json")

    plan = db_session.query(Plan).filter(Plan.id == result.plan_id).one()
    assert plan.plan_json not in (None, "", "{}")
    parsed = json.loads(plan.plan_json)
    assert parsed["format_version"] == "upload_v2"
    db_task_count = (
        db_session.query(TaskTemplate)
        .filter(TaskTemplate.plan_id == result.plan_id)
        .count()
    )
    assert len(parsed["tasks"]) == db_task_count == 2
    for task in parsed["tasks"]:
        assert task["task_id"]
