"""
Tests for workbook_to_xlsx.generate_xlsx full rebuild (MODIFY_WORKSHEET_PLAN_FINAL §4.2).

Strategy: build a representative canonical workbook_json, render it, re-open the
returned bytes with openpyxl, and assert the written values land in the correct
sheet/column. Round-trip is also verified through plan_ingest's deterministic
readers (the writers invert those readers, so a clean round-trip proves column
alignment). Resilience and boundary cases assert no exception + other sheets
still populated.

These tests are pure (no DB, no LLM): generate_xlsx operates on a dict over the
openpyxl template.
"""
import io

import openpyxl
import pytest

from app.services import plan_ingest as pi
from app.services import workbook_to_xlsx as wx


def _representative_json() -> dict:
    return {
        "personal_settings": {"name": "Jane Doe", "sex": "F", "height_cm": 165},
        "tasks": [
            {
                "pillar": "brief_today",
                "name": "Band row + Wall slide + Dead bug",
                "description": "Upper back supports posture",
                "timing": "05:50-06:10",
                "target_value": "2 rounds",
                "is_reference": False,
                "extra_metadata": {
                    "exercise_names": "Band row; Wall slide; Dead bug",
                    "progression": "+1 set",
                    "notes": "keep this block",
                },
            },
            {
                "pillar": "brief_today",
                "name": "Easy walk cool-down + Hip flexor stretch",
                "description": "Downshifts heart rate",
                "timing": "06:30-06:40",
                "target_value": "10 min",
                "is_reference": False,
                "extra_metadata": {
                    "exercise_names": "Easy walk cool-down; Hip flexor stretch",
                },
            },
            {
                "pillar": "brief_today",
                "name": "(must) Strength",
                "description": "weekly minimum note",
                "schedule": "weekly",
                "target_value": "2 sessions",
                "is_reference": False,
                "extra_metadata": {
                    "must": "true",
                    "week_1": "1 set",
                    "week_2": "2 sets",
                    "track": "log RPE",
                },
            },
            {
                "pillar": "supplements",
                "name": "Creatine monohydrate",
                "description": "ATP regeneration",
                "schedule": "daily",
                "timing": "AM with food",
                "target_value": "5 g",
                "source_key": "CRE-1",
                "is_reference": False,
                "extra_metadata": {
                    "status": "Active",
                    "category": "B",
                    "stop_rule": "Stop if kidney disease",
                },
            },
            {
                "pillar": "supplements",
                "name": "Omega-3 (EPA+DHA)",
                "description": "Cell membrane",
                "schedule": "daily",
                "timing": "AM breakfast",
                "target_value": "2 g",
                "is_reference": False,
                "extra_metadata": {"status": "Active"},
            },
        ],
        "rotation_days": [
            {
                "day_number": 1,
                "week_number": 1,
                "block_name": "Foundation strength A",
                "morning_time": "05:40-06:40",
                "warm_up_min": "10",
                "upper_back_core_min": "20",
                "secondary_min": "20",
                "cool_down_min": "10",
                "total_min": "60",
                "fits_60": "YES",
                "priority_exercises": "Band row; wall slide",
                "secondary_exercises": "Sit-to-stand",
                "week_rule": "Week 1: easy start",
                "notes": "shorten secondary first",
            },
            {
                "day_number": 2,
                "week_number": 1,
                "block_name": "Zone 2 + mobility",
                "morning_time": "05:40-06:40",
                "warm_up_min": "8",
                "total_min": "55",
                "fits_60": "YES",
                "priority_exercises": "Band pull-apart",
            },
            {
                "day_number": 3,
                "week_number": 1,
                "block_name": "Strength B",
                "total_min": "60",
            },
        ],
        "screenings": [
            {
                "pillar": "blood_markers",
                "name": "HbA1c",
                "description": "3-month glucose average",
                "frequency_months": 12,
                "target_value": "4.8-5.4",
            },
            {
                "pillar": "blood_markers",
                "name": "ApoB",
                "description": "Best lipid marker",
                "frequency_months": 12,
                "target_value": "<80",
            },
            {
                "pillar": "screenings_safety",
                "name": "Colonoscopy",
                "description": "Colorectal cancer prevention",
                "frequency_months": 120,
                "target_value": None,
            },
            {
                "pillar": "screenings_safety",
                "name": "Dental cleaning",
                "description": "Oral health",
                "frequency_months": 6,
                "target_value": None,
            },
        ],
    }


def _reopen(data: bytes) -> openpyxl.Workbook:
    return openpyxl.load_workbook(io.BytesIO(data))


# ── happy path: written cells land in the right sheet/column ───────────────


def test_personal_settings_still_stamped():
    data = wx.generate_xlsx(_representative_json())
    wb = _reopen(data)
    ws = wb["01_Personal_Settings"]
    assert ws["B6"].value == "Jane Doe"
    assert ws["B8"].value == "F"


def test_brief_today_main_table_written_correctly():
    data = wx.generate_xlsx(_representative_json())
    tasks = pi._read_brief_today_from_excel(data)
    main = [t for t in tasks if not t.name.startswith("(must)")]
    assert main, "expected actionable brief_today rows"
    first = main[0]
    # exercises split back across Exercise 1/2/3 → reader rejoins with ' + '
    assert first.name == "Band row + Wall slide + Dead bug"
    assert first.timing == "05:50-06:10"
    assert first.target_value == "2 rounds"
    assert first.description == "Upper back supports posture"
    assert first.extra_metadata is not None
    assert first.extra_metadata.get("progression") == "+1 set"


def test_brief_today_minimums_table_written_correctly():
    data = wx.generate_xlsx(_representative_json())
    tasks = pi._read_brief_today_from_excel(data)
    must = [t for t in tasks if t.name.startswith("(must)")]
    assert len(must) == 1
    m = must[0]
    assert m.name == "(must) Strength"
    assert m.target_value == "2 sessions"
    assert m.schedule == "weekly"
    assert m.extra_metadata is not None
    assert m.extra_metadata.get("week_1") == "1 set"


def test_supplements_written_and_no_template_rows_remain():
    data = wx.generate_xlsx(_representative_json())
    supps = pi._read_supplements_from_excel(data)
    names = [s.name for s in supps]
    # exactly the two supplements from the JSON — template rows fully cleared
    assert names == ["Creatine monohydrate", "Omega-3 (EPA+DHA)"]
    cre = supps[0]
    assert cre.target_value == "5 g"
    assert cre.schedule == "daily"
    assert cre.extra_metadata is not None
    assert cre.extra_metadata.get("status") == "Active"
    assert cre.source_key == "CRE-1"


def test_blood_markers_written_correctly():
    data = wx.generate_xlsx(_representative_json())
    bm = [s for s in pi._read_screenings_from_excel(data) if s.pillar == "blood_markers"]
    by_name = {s.name: s for s in bm}
    assert "HbA1c" in by_name
    assert by_name["HbA1c"].target_value == "4.8-5.4"
    assert by_name["HbA1c"].description == "3-month glucose average"
    assert "ApoB" in by_name


def test_screenings_safety_written_with_frequency_in_years():
    data = wx.generate_xlsx(_representative_json())
    ss = pi._read_screenings_from_excel(data)
    colo = next((s for s in ss if s.name == "Colonoscopy"), None)
    assert colo is not None
    # 120 months written back as 10 (years) → reader converts to 120 months
    assert colo.frequency_months == 120
    dental = next((s for s in ss if s.name == "Dental cleaning"), None)
    assert dental is not None
    # 6 months → 0.5 years → reader rounds 0.5*12 = 6
    assert dental.frequency_months == 6


def test_rotation_written_correctly():
    data = wx.generate_xlsx(_representative_json())
    days = pi._read_rotation_from_excel(data)
    by_num = {d.day_number: d for d in days}
    assert set(by_num) == {1, 2, 3}
    assert by_num[1].block_name == "Foundation strength A"
    assert by_num[1].total_min == "60"
    assert by_num[1].fits_60 == "YES"
    assert by_num[1].priority_exercises == "Band row; wall slide"
    assert by_num[2].block_name == "Zone 2 + mobility"


def test_non_data_sheet_unchanged():
    """A template-only sheet (15_Sources) must be byte-identical content-wise."""
    template = wx.get_or_copy_template()
    orig = openpyxl.load_workbook(template)
    orig_rows = list(orig["15_Sources"].iter_rows(values_only=True))

    data = wx.generate_xlsx(_representative_json())
    new_rows = list(_reopen(data)["15_Sources"].iter_rows(values_only=True))
    assert new_rows == orig_rows


# ── resilience: a missing/renamed sheet must not abort others ──────────────


def test_resilience_missing_sheet_does_not_500(tmp_path, monkeypatch):
    """Strip the supplements sheet → that sheet skipped, others still written."""
    template = wx.get_or_copy_template()
    wb = openpyxl.load_workbook(template)
    del wb["09_Supplements"]
    stripped = tmp_path / "stripped.xlsx"
    wb.save(stripped)
    monkeypatch.setattr(wx, "get_or_copy_template", lambda: stripped)

    data = wx.generate_xlsx(_representative_json())  # must not raise
    # other sheets still populated
    days = pi._read_rotation_from_excel(data)
    assert {d.day_number for d in days} == {1, 2, 3}
    bm = [s for s in pi._read_screenings_from_excel(data) if s.pillar == "blood_markers"]
    assert any(s.name == "HbA1c" for s in bm)


def test_resilience_renamed_sheet_falls_back_silently(tmp_path, monkeypatch):
    """Rename brief_today so no normalized match → skipped, rotation still ok."""
    template = wx.get_or_copy_template()
    wb = openpyxl.load_workbook(template)
    wb["02_Brief_Today"].title = "ZZ_Unmatched"
    renamed = tmp_path / "renamed.xlsx"
    wb.save(renamed)
    monkeypatch.setattr(wx, "get_or_copy_template", lambda: renamed)

    data = wx.generate_xlsx(_representative_json())  # must not raise
    days = pi._read_rotation_from_excel(data)
    assert {d.day_number for d in days} == {1, 2, 3}


def test_missing_template_raises_file_not_found(tmp_path, monkeypatch):
    """Pure-function contract: only FileNotFoundError when template absent."""
    missing = tmp_path / "nope.xlsx"
    monkeypatch.setattr(wx, "TEMPLATE_PATH", missing)
    monkeypatch.setattr(wx, "SOURCE_TEMPLATE", tmp_path / "also_nope.xlsx")
    with pytest.raises(FileNotFoundError):
        wx.generate_xlsx(_representative_json())


# ── boundary: empty arrays clear data rows, no crash ───────────────────────


def test_empty_arrays_clear_data_and_do_not_crash():
    empty = {
        "personal_settings": {"name": "Empty"},
        "tasks": [],
        "rotation_days": [],
        "screenings": [],
    }
    data = wx.generate_xlsx(empty)  # must not raise
    assert pi._read_rotation_from_excel(data) == []
    assert pi._read_supplements_from_excel(data) == []
    bt = pi._read_brief_today_from_excel(data)
    assert bt == []
    # personal settings still stamped
    assert _reopen(data)["01_Personal_Settings"]["B6"].value == "Empty"


def test_overflow_more_rotation_days_than_capacity_no_crash():
    j = _representative_json()
    j["rotation_days"] = [
        {"day_number": i, "block_name": f"Day {i}", "total_min": "60"}
        for i in range(1, 200)
    ]
    data = wx.generate_xlsx(j)  # capped + logged, must not raise
    days = pi._read_rotation_from_excel(data)
    assert len(days) >= 1  # wrote what fit
    assert all(d.block_name for d in days)


def test_missing_extra_metadata_handled():
    """Tasks lacking extra_metadata must not raise (None / missing keys)."""
    j = {
        "personal_settings": {},
        "tasks": [
            {
                "pillar": "supplements",
                "name": "Magnesium",
                "schedule": "daily",
                "target_value": "200 mg",
                "is_reference": False,
            }
        ],
        "rotation_days": [],
        "screenings": [],
    }
    data = wx.generate_xlsx(j)  # must not raise
    supps = pi._read_supplements_from_excel(data)
    assert [s.name for s in supps] == ["Magnesium"]


# ── Group B: unit tests for helper functions (no template, no DB) ───────────
# These tests use in-memory openpyxl workbooks only; they run without any
# filesystem dependency and are never affected by DB lock issues.


class TestNormalizePillar:
    """_normalize_pillar strips numeric prefix, lowercases, replaces spaces."""

    def test_strips_numeric_and_underscore_prefix(self):
        assert wx._normalize_pillar("02_Brief_Today") == "brief_today"

    def test_strips_numeric_prefix_no_underscore(self):
        assert wx._normalize_pillar("09Supplements") == "supplements"

    def test_lowercases_and_replaces_spaces(self):
        assert wx._normalize_pillar("Blood Markers") == "blood_markers"

    def test_strips_leading_whitespace(self):
        assert wx._normalize_pillar("  05_30Day_Rotation") == "30day_rotation"

    def test_already_normalized_passthrough(self):
        assert wx._normalize_pillar("brief_today") == "brief_today"


class TestFindHeaderRow:
    """_find_header_row locates the first row matching all required keywords."""

    @staticmethod
    def _make_ws(header_row: int, headers: list[str]) -> object:
        wb = openpyxl.Workbook()
        ws = wb.active
        for col, text in enumerate(headers, start=1):
            ws.cell(row=header_row, column=col).value = text
        return ws

    def test_finds_matching_row(self):
        ws = self._make_ws(2, ["Exercise 1", "Dose / Target", "Start Time"])
        assert wx._find_header_row(ws, ["exercise 1", "dose"]) == 2

    def test_returns_none_when_header_absent(self):
        ws = self._make_ws(2, ["Exercise 1", "Dose / Target"])
        assert wx._find_header_row(ws, ["completely_absent_column"]) is None

    def test_limit_prevents_finding_header_beyond_it(self):
        """Header at row 30 with limit=25 → None."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=30, column=1).value = "deep header"
        ws.cell(row=30, column=2).value = "deep column"
        assert wx._find_header_row(ws, ["deep header"], limit=25) is None

    def test_finds_header_exactly_at_limit(self):
        ws = self._make_ws(25, ["Boundary Header", "Extra"])
        assert wx._find_header_row(ws, ["boundary header"], limit=25) == 25

    def test_case_insensitive_matching(self):
        ws = self._make_ws(1, ["EXERCISE 1", "DOSE"])
        assert wx._find_header_row(ws, ["exercise 1", "dose"]) == 1

    def test_substring_match_works(self):
        """'dose' matches a cell containing 'Dose / Target'."""
        ws = self._make_ws(1, ["Dose / Target", "Name"])
        assert wx._find_header_row(ws, ["dose"]) == 1


class TestColMap:
    """_col_map returns {lowercased header text: 1-indexed column}."""

    def test_returns_column_indices(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = "Name"
        ws.cell(row=1, column=2).value = "Status"
        ws.cell(row=1, column=3).value = "Dose"
        result = wx._col_map(ws, 1)
        assert result == {"name": 1, "status": 2, "dose": 3}

    def test_ignores_none_cells(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = "Name"
        # column 2 intentionally None
        ws.cell(row=1, column=3).value = "Status"
        result = wx._col_map(ws, 1)
        assert len(result) == 2
        assert None not in result

    def test_lowercases_header_text(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = "Exercise 1"
        result = wx._col_map(ws, 1)
        assert "exercise 1" in result

    def test_first_duplicate_wins(self):
        """If header appears twice, the first column index is stored."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = "Name"
        ws.cell(row=1, column=2).value = "Name"  # duplicate
        result = wx._col_map(ws, 1)
        assert result["name"] == 1  # first occurrence kept


class TestFindCol:
    """_find_col matches keyword substrings against header dict keys."""

    def test_matches_keyword_substring(self):
        cols = {"exercise 1": 1, "dose / target": 2, "start time": 3}
        assert wx._find_col(cols, "exercise") == 1

    def test_matches_exact_key(self):
        assert wx._find_col({"dose": 2}, "dose") == 2

    def test_returns_none_when_no_match(self):
        cols = {"name": 1, "status": 2}
        assert wx._find_col(cols, "nonexistent") is None

    def test_first_keyword_takes_priority(self):
        """First keyword that matches any column is returned."""
        cols = {"dose / target": 5, "name": 1}
        # "dose" matches "dose / target" which is col 5
        assert wx._find_col(cols, "dose", "name") == 5

    def test_falls_through_to_second_keyword(self):
        """If first keyword misses, second is tried."""
        cols = {"name": 1}
        assert wx._find_col(cols, "dose", "name") == 1


class TestClearBlock:
    """_clear_block blanks the given columns for a row range, leaves others."""

    def test_blanks_cells_in_range(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        for r in range(1, 5):
            for c in range(1, 4):
                ws.cell(row=r, column=c).value = f"r{r}c{c}"

        wx._clear_block(ws, first_row=2, last_row=4, columns=[1, 2])

        for r in range(2, 5):
            assert ws.cell(row=r, column=1).value is None
            assert ws.cell(row=r, column=2).value is None
        # Column 3 untouched
        for r in range(2, 5):
            assert ws.cell(row=r, column=3).value is not None
        # Row 1 untouched
        assert ws.cell(row=1, column=1).value == "r1c1"

    def test_single_cell_range(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=3, column=2).value = "present"
        wx._clear_block(ws, first_row=3, last_row=3, columns=[2])
        assert ws.cell(row=3, column=2).value is None

    def test_empty_column_list_is_noop(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = "stay"
        wx._clear_block(ws, first_row=1, last_row=1, columns=[])
        assert ws.cell(row=1, column=1).value == "stay"

    def test_first_row_equals_last_row(self):
        """Single-row range clears only that row."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=5, column=3).value = "erase_me"
        ws.cell(row=4, column=3).value = "keep_me"
        wx._clear_block(ws, first_row=5, last_row=5, columns=[3])
        assert ws.cell(row=5, column=3).value is None
        assert ws.cell(row=4, column=3).value == "keep_me"


class TestMeta:
    """_meta extracts a key from task['extra_metadata'] dict safely."""

    def test_reads_key_from_dict(self):
        task = {"extra_metadata": {"category": "Cardio"}}
        assert wx._meta(task, "category") == "Cardio"

    def test_returns_none_for_missing_key(self):
        task = {"extra_metadata": {"status": "active"}}
        assert wx._meta(task, "category") is None

    def test_returns_none_when_metadata_is_not_dict(self):
        task = {"extra_metadata": "not_a_dict"}
        assert wx._meta(task, "category") is None

    def test_returns_none_when_metadata_key_absent(self):
        task: dict = {}
        assert wx._meta(task, "category") is None

    def test_returns_none_when_metadata_is_none(self):
        task = {"extra_metadata": None}
        assert wx._meta(task, "category") is None

    def test_returns_falsy_values_correctly(self):
        """A key present with value 0 or '' should return that value, not None."""
        task = {"extra_metadata": {"count": 0, "label": ""}}
        # _meta returns the dict value directly — 0 and "" are valid returns
        assert wx._meta(task, "count") == 0
